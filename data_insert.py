from openpyxl import load_workbook
import os
import pandas as pd
from domain import COMMON_COLS, SPEC_COLS

def first_sheet_name(xlsx_path: str) -> str:
    try:
        wb = load_workbook(xlsx_path, read_only=True)
        return wb.sheetnames[0] if wb.sheetnames else "Sheet1"
    except Exception:
        return "Sheet1"

def _norm(s: str) -> str:
    import re
    return re.sub(r"\s+", " ", str(s).strip().lower())

def _pick_col(df: pd.DataFrame, *aliases):
    norm_map = {_norm(c): c for c in df.columns}
    # exactos
    for a in aliases:
        if a and _norm(a) in norm_map:
            return norm_map[_norm(a)]
    # contains único
    for a in aliases:
        if not a:
            continue
        na = _norm(a)
        cand = [real for norm, real in norm_map.items() if na in norm or norm in na]
        if len(cand) == 1:
            return cand[0]
    return None

def _sheet_exists(xlsx_path: str, sheet_name: str) -> bool:
    try:
        wb = load_workbook(xlsx_path, read_only=True)
        return sheet_name in wb.sheetnames
    except Exception:
        return False

def append_user_to_excel(xlsx_path: str, tipo: str, row_data: dict, sheet_name: str | None):
    """
    Añade una fila al Excel en la hoja `sheet_name`.
    - Si la hoja NO existe -> la crea con cabeceras estándar.
    - Si SÍ existe -> mapea a las columnas reales y añade una fila (replace esa hoja).
    Devuelve: (ok: bool, err: str|None)
    """
    if not xlsx_path or not os.path.exists(xlsx_path):
        return False, f"No existe el Excel: {xlsx_path}"

    target_sheet = (sheet_name or "").strip() or first_sheet_name(xlsx_path)
    lat = None
    lon = None
    # row_data puede venir con "coordenadas": (lat, lon)
    if "coordenadas" in row_data and isinstance(row_data["coordenadas"], (tuple, list)) and len(row_data["coordenadas"]) == 2:
        lat, lon = row_data["coordenadas"]
    else:
        lat = row_data.get("lat")
        lon = row_data.get("lon")

    # ── Hoja NO existe → crear con columnas estándar ────────────────────────────
    if not _sheet_exists(xlsx_path, target_sheet):
        need_cols = COMMON_COLS + (SPEC_COLS.get(tipo) or [])
        new = {
            "Nombre":      row_data.get("nombre"),
            "Apellidos":   row_data.get("apellidos"),
            "Email":       row_data.get("email"),
            "Universidad": row_data.get("destino_origen"),
            "Coordenadas": (
                f"{row_data.get('coordenadas')[0]}, {row_data.get('coordenadas')[1]}"
                if isinstance(row_data.get('coordenadas'), (tuple, list)) and len(row_data['coordenadas']) == 2
                else None
            ),
        }
        if tipo == "Erasmus OUT":
            new.update({"ToR": row_data.get("tor"), "Curso": row_data.get("curso"), "ActaEquivalencias": row_data.get("acta_equivalencias")})
        elif tipo == "Erasmus IN":
            new.update({"LA": row_data.get("la"), "Horario": row_data.get("horario")})
        else:  # SICUE OUT
            new.update({"LA": row_data.get("la"), "EstadoFirmas": row_data.get("estado_firmas"), "PlanEstudios": row_data.get("plan_estudios")})

        out = pd.DataFrame([new], columns=need_cols)

        mode = "a" if os.path.exists(xlsx_path) else "w"
        try:
            with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode=mode) as w:
                out.to_excel(w, sheet_name=target_sheet, index=False)
        except PermissionError:
            return False, "El archivo está abierto en otra aplicación."
        except Exception as e:
            return False, f"Error creando la hoja '{target_sheet}': {e}"
        return True, None

    # ── Hoja SÍ existe → leer, mapear y añadir ─────────────────────────────────
    try:
        df = pd.read_excel(xlsx_path, sheet_name=target_sheet, engine="openpyxl")
    except Exception as e:
        return False, f"Error leyendo hoja '{target_sheet}': {e}"

    df.columns = [str(c).strip() for c in df.columns]
    cols_order = list(df.columns)

    # columnas comunes reales
    c_nombre    = _pick_col(df, "nombre", "Nombre")
    c_apellidos = _pick_col(df, "apellidos", "Apellidos")
    c_ap1       = _pick_col(df, "apellido1", "Apellido1")
    c_ap2       = _pick_col(df, "apellido2", "Apellido2")
    c_email     = _pick_col(df, "email", "Email")
    c_coords    = _pick_col(df, "Coordenadas", "coordenadas")
    c_lat       = _pick_col(df, "Latitud", "latitud", "Lat")
    c_lon       = _pick_col(df, "Longitud", "longitud", "Lon")

    if tipo == "Erasmus IN":
        c_univ = _pick_col(df, "Universidad Origen", "universidad origen", "Universidad", "Origen")
    else:
        c_univ = _pick_col(df, "Destino", "Universidad Destino", "universidad destino", "Universidad")

    # específicas por tipo
    if tipo == "SICUE OUT":
        c_la      = _pick_col(df, "LA", "la")
        c_gestion = _pick_col(df, "Gestion LA", "Gestión LA", "gestion la", "gestión la")
        c_estado  = _pick_col(df, "EstadoFirmas", "Estado firmas", "estado de firmas")
        c_plan    = _pick_col(df, "Enlace plan de estudios", "plan de estudios", "PlanEstudios")
    elif tipo == "Erasmus OUT":
        c_la      = _pick_col(df, "LA", "la")
        c_tor     = _pick_col(df, "ToR", "tor")
        c_curso   = _pick_col(df, "Curso", "curso")
        c_acta    = _pick_col(df, "Acta de equivalencias", "ActaEquivalencias", "acta equivalencias")
    else:  # Erasmus IN
        c_la      = _pick_col(df, "LA", "la")
        c_horario = _pick_col(df, "Horario", "horario")

    # construir nueva fila respetando columnas reales
    new_row = {c: None for c in cols_order}
    if c_nombre: new_row[c_nombre] = row_data.get("nombre")

    apes = (row_data.get("apellidos") or "").strip()
    if c_apellidos:
        new_row[c_apellidos] = apes
    else:
        parts = apes.split()
        if c_ap1: new_row[c_ap1] = parts[0] if parts else ""
        if c_ap2: new_row[c_ap2] = " ".join(parts[1:]) if len(parts) > 1 else ""

    if c_email: new_row[c_email] = row_data.get("email")
    if c_univ:  new_row[c_univ]  = row_data.get("destino_origen")

    if c_coords and (lat is not None and lon is not None):
        new_row[c_coords] = f"{lat}, {lon}"
    else:
        if c_lat: new_row[c_lat] = lat
        if c_lon: new_row[c_lon] = lon

    if tipo == "SICUE OUT":
        if c_la:      new_row[c_la]      = row_data.get("la")
        if c_gestion: new_row[c_gestion] = row_data.get("gestion_la") or row_data.get("gestion")
        if c_estado:  new_row[c_estado]  = row_data.get("estado_firmas")
        if c_plan:    new_row[c_plan]    = row_data.get("plan_estudios")
    elif tipo == "Erasmus OUT":
        if c_la:    new_row[c_la]    = row_data.get("la")
        if c_tor:   new_row[c_tor]   = row_data.get("tor")
        if c_curso: new_row[c_curso] = row_data.get("curso")
        if c_acta:  new_row[c_acta]  = row_data.get("acta_equivalencias")
    else:
        if c_la:      new_row[c_la]      = row_data.get("la")
        if c_horario: new_row[c_horario] = row_data.get("horario")

    out = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True).reindex(columns=cols_order)

    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
            out.to_excel(w, sheet_name=target_sheet, index=False)
    except TypeError:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="w") as w:
            out.to_excel(w, sheet_name=target_sheet, index=False)
    except PermissionError:
        return False, "El archivo está abierto en otra aplicación."
    except Exception as e:
        return False, f"Error guardando: {e}"

    return True, None