"""
Tests para export/stats_export.py

Funciones testeables (puras o con openpyxl/pandas inline):
  - _safe_sheet          — normaliza nombres de hoja
  - _style_title         — escribe en worksheet mock
  - _style_table         — aplica estilos en worksheet mock
  - _autosize            — calcula anchos de columna
  - _write_block         — escribe un bloque título+tabla
  - build_stats_excel    — genera bytes de Excel bien formado
"""
import pytest
import pandas as pd
from io import BytesIO
from unittest.mock import MagicMock, patch, call
from openpyxl import load_workbook

from export.stats_export import (
    _safe_sheet,
    _style_title,
    _style_table,
    _autosize,
    _write_block,
    build_stats_excel,
)


# ─────────────────────────────────────────────────────────────────────────────
# _safe_sheet
# ─────────────────────────────────────────────────────────────────────────────

class TestSafeSheet:
    def test_nombre_limpio_sin_cambios(self):
        assert _safe_sheet("Erasmus IN") == "Erasmus IN"

    def test_caracteres_prohibidos_reemplazados(self):
        result = _safe_sheet("Hoja\\/*?:[]")
        assert "\\" not in result
        assert "/" not in result
        assert "*" not in result
        assert "?" not in result
        assert ":" not in result
        assert "[" not in result
        assert "]" not in result

    def test_trunca_a_31_caracteres(self):
        largo = "A" * 50
        result = _safe_sheet(largo)
        assert len(result) <= 31

    def test_nombre_vacio_devuelve_hoja(self):
        assert _safe_sheet("") == "Hoja"

    def test_solo_espacios_devuelve_hoja(self):
        assert _safe_sheet("   ") == "Hoja"

    def test_exactamente_31_caracteres(self):
        nombre = "A" * 31
        result = _safe_sheet(nombre)
        assert result == nombre
        assert len(result) == 31

    def test_strip_espacios_laterales(self):
        result = _safe_sheet("  MiHoja  ")
        assert result == "MiHoja"

    def test_nombre_con_slash_reemplazado(self):
        result = _safe_sheet("2023/2024")
        assert "/" not in result
        assert "2023" in result
        assert "2024" in result


# ─────────────────────────────────────────────────────────────────────────────
# _style_title  (necesita un worksheet real de openpyxl)
# ─────────────────────────────────────────────────────────────────────────────

class TestStyleTitle:
    def _make_ws(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        return ws

    def test_escribe_titulo_en_celda(self):
        ws = self._make_ws()
        _style_title(ws, row=1, max_col=3, title="Mi Título")
        assert ws.cell(1, 1).value == "Mi Título"

    def test_max_col_cero_no_lanza(self):
        ws = self._make_ws()
        # max_col=0 debe tratarse como 1 (max(1, 0))
        _style_title(ws, row=1, max_col=0, title="Test")
        assert ws.cell(1, 1).value == "Test"

    def test_fuente_negrita_blanca(self):
        ws = self._make_ws()
        _style_title(ws, row=1, max_col=2, title="Header")
        font = ws.cell(1, 1).font
        assert font.bold is True
        # openpyxl almacena el color con prefijo alpha (00FFFFFF o FFFFFFFF)
        assert font.color.rgb.upper().endswith("FFFFFF")

    def test_relleno_azul(self):
        ws = self._make_ws()
        _style_title(ws, row=1, max_col=2, title="Header")
        fill = ws.cell(1, 1).fill
        # openpyxl puede usar prefijo alpha: 001F4E79 o FF1F4E79
        assert fill.fgColor.rgb.upper().endswith("1F4E79")


# ─────────────────────────────────────────────────────────────────────────────
# _style_table
# ─────────────────────────────────────────────────────────────────────────────

class TestStyleTable:
    def _make_ws_with_data(self, nrows, ncols):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for r in range(1, nrows + 2):  # +1 header, nrows data
            for c in range(1, ncols + 1):
                ws.cell(r, c, f"R{r}C{c}")
        return ws

    def test_no_lanza_con_tabla_pequena(self):
        ws = self._make_ws_with_data(3, 2)
        _style_table(ws, start_row=1, nrows=3, ncols=2)
        # Si llega aquí sin excepción, OK
        assert True

    def test_cabecera_tiene_fuente_negrita(self):
        ws = self._make_ws_with_data(2, 2)
        _style_table(ws, start_row=1, nrows=2, ncols=2)
        assert ws.cell(1, 1).font.bold is True

    def test_cuerpo_sin_negrita(self):
        ws = self._make_ws_with_data(2, 2)
        _style_table(ws, start_row=1, nrows=2, ncols=2)
        # La fila 2 (primera de datos) no debe ser negrita
        assert ws.cell(2, 1).font.bold is not True


# ─────────────────────────────────────────────────────────────────────────────
# _autosize
# ─────────────────────────────────────────────────────────────────────────────

class TestAutosize:
    def _make_ws(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        return ws

    def test_no_lanza_con_hoja_vacia(self):
        ws = self._make_ws()
        _autosize(ws)

    def test_ancho_minimo_10(self):
        ws = self._make_ws()
        ws.cell(1, 1, "X")  # valor muy corto
        _autosize(ws)
        from openpyxl.utils import get_column_letter
        letter = get_column_letter(1)
        assert ws.column_dimensions[letter].width >= 10

    def test_ancho_maximo_55(self):
        ws = self._make_ws()
        ws.cell(1, 1, "A" * 200)  # valor muy largo
        _autosize(ws)
        from openpyxl.utils import get_column_letter
        letter = get_column_letter(1)
        assert ws.column_dimensions[letter].width <= 55

    def test_ancho_proporcional_al_contenido(self):
        ws = self._make_ws()
        ws.cell(1, 1, "Corto")
        ws.cell(1, 2, "Este texto es bastante más largo que el otro")
        _autosize(ws)
        from openpyxl.utils import get_column_letter
        w1 = ws.column_dimensions[get_column_letter(1)].width
        w2 = ws.column_dimensions[get_column_letter(2)].width
        assert w2 > w1


# ─────────────────────────────────────────────────────────────────────────────
# _write_block
# ─────────────────────────────────────────────────────────────────────────────

class TestWriteBlock:
    def _make_ws(self):
        from openpyxl import Workbook
        wb = Workbook()
        return wb.active

    def test_df_vacio_escribe_sin_datos(self):
        ws = self._make_ws()
        df = pd.DataFrame()
        next_row = _write_block(ws, row=1, title="Sin datos", df=df)
        assert next_row > 1
        assert ws.cell(1, 1).value == "Sin datos"

    def test_df_none_tratado_como_vacio(self):
        ws = self._make_ws()
        next_row = _write_block(ws, row=1, title="Vacío", df=None)
        assert next_row > 1

    def test_df_normal_escribe_datos(self):
        ws = self._make_ws()
        df = pd.DataFrame({"País": ["España", "Francia"], "Total": [10, 5]})
        next_row = _write_block(ws, row=1, title="Por país", df=df)
        # La fila 1 es el título, fila 2 cabecera, filas 3-4 datos
        assert ws.cell(1, 1).value == "Por país"
        # next_row debe ser mayor que 1 + 1(cabecera) + 2(datos)
        assert next_row > 4

    def test_retorna_entero(self):
        ws = self._make_ws()
        df = pd.DataFrame({"A": [1, 2]})
        result = _write_block(ws, row=1, title="T", df=df)
        assert isinstance(result, int)

    def test_espacio_entre_bloques(self):
        ws = self._make_ws()
        df = pd.DataFrame({"A": [1]})
        r1 = _write_block(ws, row=1, title="Bloque 1", df=df)
        r2 = _write_block(ws, row=r1, title="Bloque 2", df=df)
        # El segundo bloque empieza donde acabó el primero
        assert r2 > r1


# ─────────────────────────────────────────────────────────────────────────────
# build_stats_excel
# ─────────────────────────────────────────────────────────────────────────────

class TestBuildStatsExcel:
    def test_genera_bytes_no_vacios(self):
        tables = [("Hoja1", pd.DataFrame({"A": [1, 2], "B": [3, 4]}))]
        result = build_stats_excel(tables)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_excel_valido_legible(self):
        tables = [("Resumen", pd.DataFrame({"X": [10], "Y": [20]}))]
        result = build_stats_excel(tables)
        wb = load_workbook(BytesIO(result))
        assert "Resumen" in wb.sheetnames

    def test_hoja_con_bloques(self):
        df = pd.DataFrame({"País": ["España"], "Alumnos": [5]})
        tables = [("Multi", [("Título bloque", df)])]
        result = build_stats_excel(tables)
        wb = load_workbook(BytesIO(result))
        assert "Multi" in wb.sheetnames

    def test_meta_genera_hoja_resumen(self):
        tables = []
        meta = {"Curso": "2024/2025", "Tipo": "Erasmus OUT"}
        result = build_stats_excel(tables, meta=meta)
        wb = load_workbook(BytesIO(result))
        assert "Resumen" in wb.sheetnames

    def test_export_warnings_genera_hoja_avisos(self):
        tables = []
        warnings = ["Aviso 1", "Aviso 2"]
        result = build_stats_excel(tables, export_warnings=warnings)
        wb = load_workbook(BytesIO(result))
        assert "Avisos" in wb.sheetnames

    def test_sin_meta_ni_warnings(self):
        tables = [("Datos", pd.DataFrame({"Col": ["a", "b"]}))]
        result = build_stats_excel(tables)
        wb = load_workbook(BytesIO(result))
        assert "Datos" in wb.sheetnames
        assert "Resumen" not in wb.sheetnames
        assert "Avisos" not in wb.sheetnames

    def test_tabla_con_df_vacio(self):
        tables = [("Vacía", [("Sin filas", pd.DataFrame())])]
        result = build_stats_excel(tables)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_multiples_hojas(self):
        tables = [
            ("Hoja A", pd.DataFrame({"X": [1]})),
            ("Hoja B", pd.DataFrame({"Y": [2]})),
        ]
        result = build_stats_excel(tables)
        wb = load_workbook(BytesIO(result))
        assert "Hoja A" in wb.sheetnames
        assert "Hoja B" in wb.sheetnames

    def test_nombre_hoja_con_caracteres_prohibidos(self):
        # El nombre contiene '/' que debe ser reemplazado
        tables = [("2023/2024", pd.DataFrame({"A": [1]}))]
        result = build_stats_excel(tables)
        wb = load_workbook(BytesIO(result))
        # Alguna hoja debe existir (con el nombre saneado)
        assert len(wb.sheetnames) >= 1
