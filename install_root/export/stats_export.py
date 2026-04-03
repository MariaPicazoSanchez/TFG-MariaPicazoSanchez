
from io import BytesIO
from typing import Any
import pandas as pd

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def _safe_sheet(name: str) -> str:
    _bad_chars = str.maketrans({c: "-" for c in '\\/*?:[]'})
    name = name.translate(_bad_chars).strip()
    return name[:31] or "Hoja"


def _style_title(ws, row: int, max_col: int, title: str) -> None:
    max_col = max(1, max_col)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=13, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor="1F4E79")  # azul
    ws.row_dimensions[row].height = 20


def _style_table(ws, start_row: int, nrows: int, ncols: int) -> None:
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True)

    # header
    for c in range(1, ncols + 1):
        cell = ws.cell(row=start_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # body
    for r in range(start_row + 1, start_row + nrows):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        best = 10
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            best = max(best, min(len(str(v)) + 2, 55))
        ws.column_dimensions[letter].width = best


def _write_block(ws, row: int, title: str, df: pd.DataFrame) -> int:
    if df is None:
        df = pd.DataFrame()

    ncols = max(2, len(df.columns) if not df.empty else 2)
    _style_title(ws, row=row, max_col=ncols, title=title)
    row += 1

    if df.empty:
        ws.cell(row=row, column=1, value="Sin datos").font = Font(italic=True, color="6C757D")
        return row + 3

    start_row = row
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
        row += 1

    nrows = df.shape[0] + 1
    _style_table(ws, start_row=start_row, nrows=nrows, ncols=df.shape[1])

    # filtro en cabecera
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(df.shape[1])}{start_row + df.shape[0]}"

    return row + 2  # espacio entre bloques


def build_stats_excel(
    tables: list[tuple[str, Any]],
    meta: dict[str, str] | None = None,
    export_warnings: list[str] | None = None,
) -> bytes:
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        wb = writer.book

        # Elimina la hoja por defecto "Sheet" si existe
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        if meta:
            pd.DataFrame(list(meta.items()), columns=["Campo", "Valor"]).to_excel(
                writer, sheet_name="Resumen", index=False
            )
        if export_warnings:
            pd.DataFrame({"Avisos": export_warnings}).to_excel(
                writer, sheet_name="Avisos", index=False
            )

        for sheet_name, payload in tables:
            sheet = _safe_sheet(sheet_name)

            # Caso: hoja simple (1 DF)
            if isinstance(payload, pd.DataFrame):
                payload.to_excel(writer, sheet_name=sheet, index=False)
                continue

            # Caso: hoja con bloques (varias tablas)
            blocks: list[tuple[str, pd.DataFrame]] = payload
            ws = wb.create_sheet(title=sheet)
            row = 1
            for title, df in blocks:
                row = _write_block(ws, row=row, title=title, df=df)
            _autosize(ws)

    return out.getvalue()
