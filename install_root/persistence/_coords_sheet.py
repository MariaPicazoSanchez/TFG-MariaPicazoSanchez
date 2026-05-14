"""
Mantiene la hoja "Coordenadas" del Excel sincronizada con las universidades
que aparecen en los alumnos.

Cuando se guarda/edita un alumno cuya universidad aún no figura en la hoja
"Coordenadas" del Excel correspondiente, esta utilidad la añade
geocodificándola con Nominatim. Las coordenadas auto-generadas se marcan con
el sufijo " (auto)" para que el coordinador sepa que las puede revisar y
sobrescribir manualmente.
"""

from __future__ import annotations

import logging
from typing import Optional

from openpyxl import load_workbook

logger = logging.getLogger("movilidad_persistence")

AUTO_SUFFIX = " (auto)"

# Geocoder independiente de Streamlit (no usa st.session_state) para poder
# llamarse también desde el endpoint Flask de edición.
try:
    from geopy.geocoders import Nominatim
    from geopy.extra.rate_limiter import RateLimiter
    _GEOCODER = Nominatim(user_agent="tfg-movilidad-esii")
    _GEOCODE = RateLimiter(_GEOCODER.geocode, min_delay_seconds=1)
except Exception:
    _GEOCODE = None

_PROCESS_CACHE: dict[str, Optional[tuple[float, float]]] = {}


def _geocode(q: str) -> Optional[tuple[float, float]]:
    if not q or not q.strip() or _GEOCODE is None:
        return None
    qn = q.strip().lower()
    if qn in _PROCESS_CACHE:
        return _PROCESS_CACHE[qn]
    try:
        loc = _GEOCODE(q, addressdetails=False, timeout=10)
    except Exception as e:
        logger.debug("[coords-sheet] geocode '%s' falló: %s", q, e)
        _PROCESS_CACHE[qn] = None
        return None
    if loc:
        res = (float(loc.latitude), float(loc.longitude))
        _PROCESS_CACHE[qn] = res
        return res
    _PROCESS_CACHE[qn] = None
    return None


def _norm(s) -> str:
    return " ".join(str(s or "").strip().lower().split())


def _resolve_columns(programa: str) -> tuple[int, int, int]:
    """Devuelve (col_uni_1b, col_pais_1b, col_coords_1b) según el programa.

    - Erasmus OUT: col0=Universidad, col1=País, col2=Coordenadas
    - Erasmus IN / SICUE OUT: col0=País, col1=Universidad, col2=Coordenadas
    """
    if (programa or "").strip().lower() == "erasmus out":
        return 1, 2, 3
    return 2, 1, 3


def _find_coords_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower() == "coordenadas":
            return wb[name]
    return None


def ensure_university_in_coords_sheet(
    xlsx_path: str,
    programa: str,
    universidad: str,
    pais: str = "",
) -> bool:
    """
    Si `universidad` no está en la hoja "Coordenadas" del Excel, geocodifica
    con Nominatim y añade una fila con las coordenadas marcadas con
    " (auto)". Devuelve True si se añadió fila, False en otro caso (ya
    existía, no se pudo geocodificar, error al abrir el fichero, etc.).
    """
    uni = (universidad or "").strip()
    if not uni:
        return False

    col_uni, col_pais, col_coords = _resolve_columns(programa)
    wb = None
    try:
        wb = load_workbook(xlsx_path)
        ws = _find_coords_sheet(wb)
        if ws is None:
            return False

        uni_norm = _norm(uni)
        last_row = ws.max_row or 1
        first_empty = None
        for r in range(1, last_row + 1):
            existing = ws.cell(row=r, column=col_uni).value
            if existing is None or str(existing).strip() == "":
                if first_empty is None:
                    first_empty = r
                continue
            if _norm(existing) == uni_norm:
                return False  # ya está

        coords = _geocode(uni)
        if coords is None and pais:
            coords = _geocode(f"{uni}, {pais}")
        if coords is None:
            logger.info("[coords-sheet] No se pudo geocodificar '%s'", uni)
            return False

        lat, lon = coords
        coords_str = f"{lat}, {lon}{AUTO_SUFFIX}"
        insert_row = first_empty if first_empty is not None else last_row + 1

        ws.cell(row=insert_row, column=col_uni).value = uni
        if pais:
            ws.cell(row=insert_row, column=col_pais).value = pais
        ws.cell(row=insert_row, column=col_coords).value = coords_str

        wb.save(xlsx_path)
        logger.info(
            "[coords-sheet] '%s' añadida en %s fila %d con %s",
            uni, xlsx_path, insert_row, coords_str,
        )
        return True

    except PermissionError:
        logger.warning("[coords-sheet] '%s' abierto en otra app, no se pudo actualizar.", xlsx_path)
        return False
    except Exception as e:
        logger.warning("[coords-sheet] Error añadiendo '%s': %s", universidad, e)
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


# ─────────────────────────────────────────────────────────────────────────────
# SICUE OUT: coordenadas en una columna del propio curso
# ─────────────────────────────────────────────────────────────────────────────

def resolve_sicue_coords_for_universidad(
    xlsx_path: str,
    universidad: str,
    *,
    exclude_sheet: str | None = None,
    exclude_row: int | None = None,
) -> str | None:
    """
    Resuelve el valor de la columna "Coordenadas" para una universidad SICUE
    OUT. La hoja SICUE OUT NO tiene una hoja "Coordenadas" separada: las
    coordenadas viven en una columna de cada hoja de curso académico.

    Estrategia:
      1) Buscar otra fila (en cualquier hoja del workbook) con la misma
         universidad y devolver su valor de Coordenadas si lo tiene.
      2) Si no hay match, geocodificar con Nominatim y devolver el valor con
         el sufijo " (auto)".
      3) Si tampoco se puede geocodificar, devolver None.
    """
    uni_norm = _norm(universidad)
    if not uni_norm:
        return None

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        logger.warning("[coords-sicue] no se pudo abrir %s: %s", xlsx_path, e)
        return None

    try:
        for sheet in wb.sheetnames:
            if sheet.strip().lower() == "coordenadas":
                continue
            ws = wb[sheet]
            # Cabecera en fila 1
            headers: dict[str, int] = {}
            for c in range(1, (ws.max_column or 0) + 1):
                v = ws.cell(row=1, column=c).value
                if v:
                    headers[_norm(v)] = c
            c_dest = headers.get("destino")
            c_coords = headers.get("coordenadas")
            if not c_dest or not c_coords:
                continue
            for r in range(2, (ws.max_row or 1) + 1):
                if exclude_sheet == sheet and exclude_row == r:
                    continue
                v = ws.cell(row=r, column=c_dest).value
                if v and _norm(v) == uni_norm:
                    coords_val = ws.cell(row=r, column=c_coords).value
                    if coords_val and str(coords_val).strip():
                        return str(coords_val).strip()
    finally:
        try:
            wb.close()
        except Exception:
            pass

    # Sin match: geocodificar
    coords = _geocode(universidad)
    if coords is None:
        return None
    lat, lon = coords
    return f"{lat}, {lon}{AUTO_SUFFIX}"
