"""
Utilidades puras reutilizadas por el módulo de inserción.

Exporta:
  - first_sheet_name  — primera hoja de un workbook
  - _norm             — normalización de texto para comparar columnas
  - _pick_col         — localiza columna en DataFrame por alias
  - _sheet_exists     — comprueba si una hoja existe en un xlsx
"""

from __future__ import annotations

import logging
import re

from openpyxl import load_workbook
import pandas as pd

logger = logging.getLogger("movilidad_persistence")


def first_sheet_name(xlsx_path: str) -> str:
    try:
        wb = load_workbook(xlsx_path, read_only=True)
        return wb.sheetnames[0] if wb.sheetnames else "Sheet1"
    except Exception:
        return "Sheet1"


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _pick_col(df: pd.DataFrame, *aliases) -> str | None:
    norm_map = {_norm(c): c for c in df.columns}

    for a in aliases:
        if a and _norm(a) in norm_map:
            return norm_map[_norm(a)]

    for a in aliases:
        if not a:
            continue
        na = _norm(a)
        cand = [real for norm, real in norm_map.items() if na in norm or norm in na]
        if len(cand) == 1:
            return cand[0]

    for a in aliases:
        if not a:
            continue
        na = _norm(a)
        for norm, real in norm_map.items():
            if na in norm:
                return real

    if aliases:
        logger.warning("No se encontró columna para alias: %s en columnas: %s", aliases, list(df.columns))
    return None


def _sheet_exists(xlsx_path: str, sheet_name: str) -> bool:
    try:
        wb = load_workbook(xlsx_path, read_only=True)
        return sheet_name in wb.sheetnames
    except Exception:
        return False
