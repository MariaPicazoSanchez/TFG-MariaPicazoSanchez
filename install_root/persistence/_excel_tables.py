"""
Detección y navegación de tablas dentro de workbooks openpyxl.

Exporta:
  - TableInfo          — dataclass con metadatos de una tabla
  - _find_table_in_workbook    — busca la primera tabla que encaje
  - _iter_all_tables_in_workbook — generador de todas las tablas
  - _build_header_maps_from_ws  — construye mapas de cabecera
  - _find_col_in_ws_by_aliases  — localiza columna por alias
  - _norm_header / normalize_str — normalización de texto
  - Constantes STUDENTS_HEADER_ALIASES, MATERIAS_HEADER_ALIASES, ...
"""


import logging
import unicodedata
from dataclasses import dataclass
from typing import Any, Iterable, Optional

from openpyxl import load_workbook

logger = logging.getLogger("movilidad_persistence")

# ─────────────────────────────────────────────────────────────────────────────
# Alias de cabeceras para detectar tablas
# ─────────────────────────────────────────────────────────────────────────────

STUDENTS_HEADER_ALIASES: dict[str, set] = {
    "nombre":             {"nombre", "nombre completo", "estudiante"},
    "apellido1":          {"apellido1", "apellido 1", "primer apellido", "apellidos"},
    "apellido2":          {"apellido2", "apellido 2", "segundo apellido"},
    "email":              {"email", "e-mail", "correo", "correo electronico", "correo electrónico"},
    "pais":               {"pais", "país"},
    "cuatrimestre":       {"cuatrimestre", "cuat", "cuatri"},
    "universidad_origen": {"universidad de origen", "universidad origen", "origen", "destino", "universidad"},
    "coordenadas":        {"coordenadas", "coords"},
    "ciudad":             {"ciudad"},
}
STUDENTS_REQUIRED: set[str] = {"nombre"}

MATERIAS_HEADER_ALIASES: dict[str, set] = {
    "asignatura":         {"asignatura"},
    "estudiante":         {"estudiante", "estudiantes", "alumno", "alumnos"},
    "origen":             {"origen"},
    "universidad_origen": {"universidad de origen", "universidad origen", "centro", "universidadorigen"},
    "cuat":               {"cuat", "cuatrimestre", "cuatri"},
    "firmado":            {"firmado", "firma"},
    "link_la":            {"link la", "linkla", "la", "learning agreement", "link_la"},
}
MATERIAS_REQUIRED: set[str] = {"asignatura", "estudiante"}


# ─────────────────────────────────────────────────────────────────────────────
# Normalización de texto
# ─────────────────────────────────────────────────────────────────────────────

def _norm_header(s: Any) -> str:
    s = str(s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def normalize_str(s: Any) -> str:
    return _norm_header(s)


# ─────────────────────────────────────────────────────────────────────────────
# TableInfo
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class TableInfo:
    sheet_name: str
    header_row: int       # 1-based
    data_start: int       # 1-based
    data_end:   int       # 1-based inclusive
    cols:       dict[str, int]  # clave canónica -> columna 1-based


# ─────────────────────────────────────────────────────────────────────────────
# Helpers de hoja y fila
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_priority(sheet_name: str) -> tuple:
    n = _norm_header(sheet_name)
    return (0 if n == "curso" else 1, n)


def _iter_sheets_preferred(wb) -> Iterable:
    return sorted(wb.worksheets, key=lambda ws: _sheet_priority(ws.title))


def _row_is_empty_ws(ws, row_num: int, max_col: Optional[int] = None) -> bool:
    max_col = max_col or ws.max_column
    for c in range(1, max_col + 1):
        v = ws.cell(row=row_num, column=c).value
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True


def _match_header_row(
    ws, row_num: int,
    aliases_map: dict[str, set],
    required: set,
    extra_min_matches: int = 0,
    extras_pool: Optional[set] = None,
) -> Optional[dict[str, int]]:
    found: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=row_num, column=c).value
        n = _norm_header(v)
        if not n:
            continue
        for canon, aliases in aliases_map.items():
            if n in aliases and canon not in found:
                found[canon] = c

    if not required.issubset(found.keys()):
        return None
    if extra_min_matches and extras_pool:
        if len(extras_pool.intersection(found.keys())) < extra_min_matches:
            return None
    return found


def _scan_table_bounds(ws, header_row: int, aliases_map, required, extra_min_matches, extras_pool) -> tuple[int, int]:
    """Devuelve (data_start, data_end) para una cabecera encontrada en header_row."""
    max_r   = ws.max_row
    max_c   = ws.max_column
    data_start = header_row + 1
    data_end   = header_row
    rr = data_start
    while rr <= max_r:
        if _row_is_empty_ws(ws, rr, max_c):
            break
        if _match_header_row(ws, rr, aliases_map, required, extra_min_matches, extras_pool):
            break
        non_empty = [
            ws.cell(row=rr, column=c).value
            for c in range(1, max_c + 1)
            if ws.cell(row=rr, column=c).value not in (None, "")
        ]
        if len(non_empty) == 1:
            v = non_empty[0]
            if isinstance(v, (int, float)) or (isinstance(v, str) and v.strip().isdigit()):
                break
        data_end = rr
        rr += 1
    return data_start, data_end


# ─────────────────────────────────────────────────────────────────────────────
# Búsqueda de tablas
# ─────────────────────────────────────────────────────────────────────────────

def _iter_all_tables_in_workbook(
    wb,
    aliases_map: dict[str, set],
    required: set,
    extra_min_matches: int = 0,
    extras_pool: Optional[set] = None,
):
    """Generador: devuelve TableInfo de todas las tablas en todas las hojas."""
    for ws in wb.worksheets:
        max_r = ws.max_row
        max_c = ws.max_column
        if not max_r or not max_c:
            continue
        r = 1
        while r <= max_r:
            cols = _match_header_row(ws, r, aliases_map, required, extra_min_matches, extras_pool)
            if not cols:
                r += 1
                continue
            data_start, data_end = _scan_table_bounds(ws, r, aliases_map, required, extra_min_matches, extras_pool)
            yield TableInfo(sheet_name=ws.title, header_row=r,
                            data_start=data_start, data_end=data_end, cols=cols)
            r = data_end + 1


def _find_table_in_workbook(
    excel_path: str,
    aliases_map: dict[str, set],
    required: set,
    extra_min_matches: int = 0,
    extras_pool: Optional[set] = None,
    target_sheet: str = "",
) -> Optional[TableInfo]:
    wb = load_workbook(excel_path)
    try:
        if target_sheet and target_sheet in wb.sheetnames:
            sheets_to_search = [wb[target_sheet]]
        else:
            sheets_to_search = _iter_sheets_preferred(wb)

        for ws in sheets_to_search:
            if not ws.max_row or not ws.max_column:
                continue
            r = 1
            while r <= ws.max_row:
                cols = _match_header_row(ws, r, aliases_map, required, extra_min_matches, extras_pool)
                if not cols:
                    r += 1
                    continue
                data_start, data_end = _scan_table_bounds(
                    ws, r, aliases_map, required, extra_min_matches, extras_pool
                )
                info = TableInfo(sheet_name=ws.title, header_row=r,
                                 data_start=data_start, data_end=data_end, cols=cols)
                logger.debug(
                    "[detect] Tabla en hoja='%s' header=%d datos=%d..%d cols=%s",
                    info.sheet_name, info.header_row, info.data_start, info.data_end, list(info.cols.keys()),
                )
                return info
        return None
    finally:
        wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# Helpers de cabecera
# ─────────────────────────────────────────────────────────────────────────────

def _build_header_maps_from_ws(ws, header_row_idx_0based: int):
    """Devuelve (norm_to_col, raw_headers) para una fila de cabecera."""
    row_excel = header_row_idx_0based + 1
    norm_to_col: dict[str, int] = {}
    raw_headers: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=row_excel, column=c).value
        if v is None:
            continue
        raw = str(v).strip()
        if not raw:
            continue
        raw_headers[c] = raw
        n = _norm_header(raw)
        if n and n not in norm_to_col:
            norm_to_col[n] = c
    return norm_to_col, raw_headers


def _find_col_in_ws_by_aliases(norm_to_col_1based: dict[str, int], aliases: list[str]) -> Optional[int]:
    for a in aliases:
        n = _norm_header(a)
        if n in norm_to_col_1based:
            return norm_to_col_1based[n]
    return None
