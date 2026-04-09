import logging
import os
import unicodedata

import openpyxl as _openpyxl
import pandas as pd

logger = logging.getLogger("movilidad_persistence")


# ─────────────────────────────────────────────────────────────────────────────
# Normalización
# ─────────────────────────────────────────────────────────────────────────────

def _norm(s) -> str:
    if pd.isna(s):
        return ""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s.replace(" ", "")


def _normalize_sheet(s: str) -> str:
    return str(s).strip().replace(" ", "").replace("-", "").replace("/", "").lower()


def _safe_int(val):
    try:
        if pd.isna(val):
            return None
        s = str(val).strip().replace(",", ".")
        if s == "" or s.lower() in ("nan", "none"):
            return None
        return int(float(s))
    except Exception:
        return None


def _cell(row, idx: int) -> str:
    """Valor de la celda idx en row, o '' si fuera de rango / NaN."""
    if idx < 0 or idx >= len(row):
        return ""
    v = row.iloc[idx]
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    return str(v).strip()


def _es_numerico(s: str) -> bool:
    try:
        float(str(s).strip().replace(",", "."))
        return True
    except (ValueError, TypeError):
        return False


# ─────────────────────────────────────────────────────────────────────────────
# Alias de cabeceras para tabla Materias IN
# ─────────────────────────────────────────────────────────────────────────────

HEADER_ALIASES = {
    "asignatura":        {"asignatura", "asignaturas", "materia", "materias",
                          "nombreasignatura", "nombreasignaturas"},
    "estudiante":        {"estudiante", "estudiantes", "alumno", "alumnos",
                          "nombre", "nombrealumno", "nombreestudiante",
                          "nombreyapellidos", "apellidosynombre"},
    "origen":            {"origen", "paisorigen", "pais"},
    "universidadorigen": {"universidadorigen", "universidaddeorigen", "univorigen",
                          "univ.deorigen", "centro", "centroorigen", "centroorigen"},
    "cuat":              {"cuat", "cuatrimestre", "cuatri", "semestre", "periodo"},
    "firmado":           {"firmado", "firma"},
    "la":                {"la", "learningagreement", "acuerdoaprendizaje"},
}
REQUIRED_KEYS = {"asignatura", "estudiante"}


# ─────────────────────────────────────────────────────────────────────────────
# Detección de tablas Materias IN
# ─────────────────────────────────────────────────────────────────────────────

def _match_header_row(row_values) -> dict | None:
    """
    Intenta mapear una fila como cabecera de Materias IN.
    Devuelve {clave_canonica: indice_columna} o None.

    Solo requiere 'asignatura' + 'estudiante'. Columnas extra (origen,
    cuatrimestre, etc.) se usan si están presentes pero no son obligatorias.
    """
    found: dict[str, int] = {}
    normalized_cells = [_norm(v) for v in row_values]

    for idx, cell in enumerate(normalized_cells):
        if not cell:
            continue
        for canon_key, aliases in HEADER_ALIASES.items():
            if cell in aliases and canon_key not in found:
                found[canon_key] = idx

    if not REQUIRED_KEYS.issubset(found.keys()):
        return None

    return found


def _is_separator_or_empty_row(row_values) -> bool:
    return all(_norm(v) == "" for v in row_values)


def _extract_block_rows(df_raw, header_map: dict, start_row: int) -> list[dict]:
    """
    Extrae las filas de datos bajo una cabecera detectada hasta separador u otra cabecera.
    Devuelve (rows_data, next_row).
    """
    n_rows = len(df_raw)
    rows_data: list[dict] = []
    c_asig = header_map.get("asignatura")
    c_est  = header_map.get("estudiante")

    j = start_row
    while j < n_rows:
        current_vals = df_raw.iloc[j].tolist()
        logger.debug("[DEBUG]   Fila datos %d: %s", j, current_vals)

        if _is_separator_or_empty_row(current_vals):
            logger.debug("[DEBUG]   Fila %d vacía/separador. Fin bloque.", j)
            break

        key_asig = _norm(current_vals[c_asig]) if c_asig is not None and c_asig < len(current_vals) else ""
        key_est  = _norm(current_vals[c_est])  if c_est  is not None and c_est  < len(current_vals) else ""
        if not key_asig and not key_est:
            logger.debug("[DEBUG]   Fila %d sin asignatura ni estudiante. Fin bloque.", j)
            break

        if _match_header_row(current_vals) is not None:
            logger.debug("[DEBUG]   Fila %d parece otra cabecera. Fin bloque.", j)
            break

        record = {
            key: df_raw.iat[j, col_idx] if col_idx < df_raw.shape[1] else None
            for key, col_idx in header_map.items()
        }
        record["_sheet_name"] = ""  # se rellena en el llamador
        logger.debug("[DEBUG]   Registro extraído: %s", record)
        rows_data.append(record)
        j += 1

    return rows_data, j


def _block_is_valid(rows_data: list[dict], sheet_name: str, header_row: int) -> bool:
    """Descarta bloques donde todos los 'estudiante' son numéricos (falso positivo)."""
    est_vals = [str(r.get("estudiante") or "").strip() for r in rows_data]
    est_nonempty = [v for v in est_vals if v and v.lower() not in ("nan", "none", "")]
    if est_nonempty and all(_es_numerico(v) for v in est_nonempty):
        logger.debug(
            "[DEBUG] Bloque descartado en hoja '%s' fila %d: todos 'estudiante' numéricos (%s)",
            sheet_name, header_row, est_nonempty[:3],
        )
        return False
    return True


# ─────────────────────────────────────────────────────────────────────────────
# Carga de Materias IN
# ─────────────────────────────────────────────────────────────────────────────

def load_materias_in(config) -> pd.DataFrame:
    """
    Lee el Excel de 'Erasmus IN' y extrae todas las tablas de Materias IN
    (aunque haya otras tablas en el mismo archivo o en distintas hojas).
    """
    ruta = config.get("Erasmus IN")
    if not ruta or not os.path.exists(ruta):
        return pd.DataFrame()

    try:
        xls = pd.read_excel(ruta, sheet_name=None, header=None, dtype=str)
        bloques: list[pd.DataFrame] = []
        hojas = list(xls.keys())
        logger.warning("[MATERIAS-IN] Hojas encontradas en el Excel: %s", hojas)

        for sheet_name, df_raw in xls.items():
            if df_raw is None or df_raw.empty:
                continue
            if df_raw.dropna(how="all").empty:
                logger.debug("[MATERIAS-IN] Hoja '%s' vacía, ignorada.", sheet_name)
                continue

            logger.warning("[MATERIAS-IN] Buscando tabla en hoja '%s' (%d filas)...", sheet_name, len(df_raw))

            i = 0
            n_rows = len(df_raw)
            last_header_map: dict | None = None  # último header válido de esta hoja

            while i < n_rows:
                row_vals = df_raw.iloc[i].tolist()

                # Saltar filas vacías sin perder el header anterior
                if _is_separator_or_empty_row(row_vals):
                    i += 1
                    continue

                header_map = _match_header_row(row_vals)

                if header_map is not None:
                    # Nueva cabecera explícita encontrada
                    last_header_map = header_map
                    logger.warning(
                        "[MATERIAS-IN] ✓ Cabecera encontrada en hoja '%s', fila %d: mapa=%s",
                        sheet_name, i, header_map,
                    )
                    rows_data, next_row = _extract_block_rows(df_raw, header_map, i + 1)
                    header_row_for_log = i
                    i = next_row + 1

                elif last_header_map is not None:
                    # Fila con datos pero sin nueva cabecera → continuar con el
                    # último header conocido (grupos separados por filas vacías)
                    logger.debug(
                        "[MATERIAS-IN]   Continuando bloque desde fila %d con header previo",
                        i,
                    )
                    rows_data, next_row = _extract_block_rows(df_raw, last_header_map, i)
                    header_row_for_log = i
                    i = next_row + 1

                else:
                    # Datos antes de cualquier cabecera: ignorar
                    i += 1
                    continue

                if rows_data:
                    for r in rows_data:
                        r["_sheet_name"] = sheet_name
                    if _block_is_valid(rows_data, sheet_name, header_row_for_log):
                        bloques.append(pd.DataFrame(rows_data))
                        logger.warning(
                            "[MATERIAS-IN] ✓ Bloque válido en hoja '%s': %d filas de datos",
                            sheet_name, len(rows_data),
                        )
                    else:
                        logger.warning(
                            "[MATERIAS-IN] ✗ Bloque descartado (valores numéricos) "
                            "en hoja '%s', fila %d",
                            sheet_name, header_row_for_log,
                        )
                else:
                    logger.debug(
                        "[MATERIAS-IN]   Sin filas de datos desde fila %d en hoja '%s'",
                        header_row_for_log, sheet_name,
                    )

        if not bloques:
            logger.warning(
                "[MATERIAS-IN] ✗ No se encontró ninguna tabla de Materias IN. "
                "Comprueba que el Excel tenga columnas 'Estudiante' y 'Asignatura' "
                "(o equivalentes) en la misma fila de cabecera."
            )
            return pd.DataFrame()

        df = pd.concat(bloques, ignore_index=True)

        columnas = ["asignatura", "estudiante", "origen", "universidadorigen", "cuat", "firmado", "la", "_sheet_name"]
        for c in columnas:
            if c not in df.columns:
                df[c] = None
        df = df[columnas].rename(columns={
            "asignatura":        "Asignatura",
            "estudiante":        "Estudiante",
            "origen":            "Origen",
            "universidadorigen": "UniversidadOrigen",
            "cuat":              "Cuat",
            "firmado":           "Firmado",
            "la":                "LA",
            "_sheet_name":       "SheetName",
        })

        for c in df.columns:
            df[c] = df[c].astype(str).str.strip()
            df.loc[df[c].isin(["", "nan", "None"]), c] = pd.NA

        df = df.dropna(subset=["Estudiante", "Asignatura"]).reset_index(drop=True)

        mask_num = df["Estudiante"].apply(_es_numerico)
        if mask_num.any():
            logger.warning(
                "[MATERIAS-IN] Descartando %d filas con Estudiante numérico: %s",
                mask_num.sum(), df.loc[mask_num, "Estudiante"].tolist()[:5],
            )
        df = df[~mask_num].reset_index(drop=True)

        logger.warning(
            "[MATERIAS-IN] ✓ Carga completada: %d filas | %d alumnos únicos",
            len(df),
            df["Estudiante"].nunique(),
        )
        logger.warning(
            "[MATERIAS-IN]   Muestra de alumnos: %s",
            df["Estudiante"].dropna().unique()[:5].tolist(),
        )
        return df

    except Exception as e:
        logger.warning("[MATERIAS-IN] ✗ Excepción leyendo materias: %s", e, exc_info=True)
        return pd.DataFrame()


def build_materias_in_por_estudiante(df_materias: pd.DataFrame) -> dict:
    """Devuelve {nombre_estudiante: [{datos asignatura}, ...]}."""
    if df_materias is None or df_materias.empty:
        return {}

    materias_por_est: dict[str, list] = {}
    for row in df_materias.itertuples(index=False, name="Row"):
        est = str(getattr(row, "Estudiante", "") or "").strip()
        if not est:
            continue
        materias_por_est.setdefault(est, []).append({
            "asignatura": getattr(row, "Asignatura", None),
            "cuat":       getattr(row, "Cuat", None),
            "firmado":    getattr(row, "Firmado", None),
            "origen":     getattr(row, "Origen", None),
            "centro":     getattr(row, "UniversidadOrigen", None),
            "la":         getattr(row, "LA", None),
            "sheet_name": getattr(row, "SheetName", None) or "",
        })
    return materias_por_est


def get_materias_in_por_estudiante(config) -> dict:
    return build_materias_in_por_estudiante(load_materias_in(config))


# ─────────────────────────────────────────────────────────────────────────────
# Catálogo de asignaturas ofertadas
# ─────────────────────────────────────────────────────────────────────────────

def _find_catalog_block(row_values) -> dict | None:
    """
    Busca en una fila la subcabecera del catálogo de asignaturas.
    Elige el candidato 'asignatura' con mayor puntuación (matr+cupo = catálogo real).
    Devuelve dict con índices de columna o None.
    """
    normalized = [_norm(v) for v in row_values]

    asig_candidates = [
        idx for idx, cell in enumerate(normalized)
        if cell in ("asignaturas", "asignatura", "materias", "materia")
    ]
    if not asig_candidates:
        return None

    best = None
    best_score = -1

    for asig_idx in asig_candidates:
        indices = {"asig": asig_idx, "cuat": -1, "matr": -1, "cupo": -1}
        for idx in range(asig_idx + 1, min(asig_idx + 10, len(normalized))):
            cell = normalized[idx]
            if cell in ("cuat", "cuatrimestre", "periodo", "semestre"):
                indices["cuat"] = idx
            elif cell in ("matriculados", "matricula", "nmatric", "inscritos", "matr"):
                indices["matr"] = idx
            elif cell in ("cupo", "plazas", "capacidad", "cupos", "aforo"):
                indices["cupo"] = idx

        has_matr = indices["matr"] != -1
        has_cupo = indices["cupo"] != -1
        has_cuat = indices["cuat"] != -1
        score = (has_matr and has_cupo) * 4 + has_matr * 2 + has_cupo * 2 + has_cuat

        if score > best_score and (has_matr or has_cupo):
            best_score = score
            best = indices

    return best


def _build_countif_cache(df_raw, header_row: int, alumnos_asig_idx: int) -> dict[str, int]:
    """Cuenta ocurrencias de cada asignatura en la tabla de alumnos (fallback COUNTIF)."""
    cache: dict[str, int] = {}
    n_rows = len(df_raw)
    for r in range(header_row + 1, n_rows):
        val = _cell(df_raw.iloc[r], alumnos_asig_idx)
        if val and val.lower() not in ("nan", "none"):
            cache[val] = cache.get(val, 0) + 1
    return cache


def _read_catalog_entry(
    row, j: int, indices: dict, ws_data, countif_cache: dict
) -> dict:
    """Construye un dict de asignatura a partir de una fila del catálogo."""
    asig_idx = indices["asig"]
    cuat_idx = indices["cuat"]
    matr_idx = indices["matr"]
    cupo_idx = indices["cupo"]

    asig = _cell(row, asig_idx)

    cuat = ""
    if cuat_idx != -1:
        c_val = _cell(row, cuat_idx)
        if c_val and c_val.lower() not in ("nan", "none"):
            cuat = c_val.split(".")[0] if "." in c_val else c_val

    matr_val = None
    if matr_idx != -1:
        if ws_data is not None:
            try:
                matr_val = _safe_int(ws_data.cell(row=j + 1, column=matr_idx + 1).value)
            except Exception:
                pass
        if matr_val is None:
            matr_val = _safe_int(row.iloc[matr_idx])
    if matr_val is None:
        matr_val = countif_cache.get(asig, 0)

    cupo_val = None
    if cupo_idx != -1:
        if ws_data is not None:
            try:
                cupo_val = _safe_int(ws_data.cell(row=j + 1, column=cupo_idx + 1).value)
            except Exception:
                pass
        if cupo_val is None:
            cupo_val = _safe_int(row.iloc[cupo_idx])

    return {"asignatura": asig, "cuat": cuat, "matriculados": matr_val, "cupo": cupo_val}


def _catalog_sort_key(r: dict) -> tuple:
    c = r.get("cuat", "")
    if c in ("1", "1.0"): return (0, r["asignatura"])
    if c in ("2", "2.0"): return (1, r["asignatura"])
    return (2, r["asignatura"])


def get_asignaturas_catalog(config, sheet_name: str | None = None) -> list[dict]:
    """
    Lee el catálogo de asignaturas ofertadas desde el Excel de Erasmus IN.
    Si se pasa sheet_name, solo busca en esa hoja.
    Devuelve lista de dicts [{asignatura, cuat, matriculados, cupo}] ordenados por cuat.
    """
    ruta = config.get("Erasmus IN") or ""
    if not ruta or not os.path.exists(ruta):
        return []

    try:
        xls = pd.read_excel(ruta, sheet_name=None, header=None, dtype=str)

        # Segunda lectura data_only para resolver fórmulas cacheadas (COUNTIF, etc.)
        try:
            _wb_data = _openpyxl.load_workbook(ruta, data_only=True, read_only=True)
            _ws_data = {ws.title: ws for ws in _wb_data.worksheets}
        except Exception:
            _wb_data = None
            _ws_data = {}

        # Seleccionar hojas a iterar
        sheet_candidates = list(xls.keys())
        sheet_to_use = None
        if sheet_name:
            norm_target = _normalize_sheet(sheet_name)
            sheet_to_use = next(
                (s for s in sheet_candidates if _normalize_sheet(s) == norm_target), None
            ) or next(
                (s for s in sheet_candidates if norm_target in _normalize_sheet(s)), None
            )
        sheets_iter = [sheet_to_use] if sheet_to_use else sheet_candidates

        rows: list[dict] = []

        for sname in sheets_iter:
            df_raw = xls[sname]
            if df_raw is None or df_raw.empty:
                continue

            ws_data = _ws_data.get(sname)
            n_rows  = len(df_raw)
            i = 0

            while i < n_rows:
                row_vals = df_raw.iloc[i].tolist()
                indices  = _find_catalog_block(row_vals)
                if indices is None:
                    i += 1
                    continue

                asig_idx = indices["asig"]
                logger.debug(
                    "[catalog] Cabecera encontrada en hoja '%s' fila %d: "
                    "asig=%d matr=%d cuat=%d cupo=%d",
                    sname, i, asig_idx, indices["matr"], indices["cuat"], indices["cupo"],
                )

                # Fallback COUNTIF manual: buscar otra columna 'asignatura' en la misma fila
                normalized_header = [_norm(v) for v in row_vals]
                alumnos_asig_idx = next(
                    (idx for idx, cell in enumerate(normalized_header)
                     if cell in ("asignatura", "asignaturas") and idx != asig_idx),
                    None,
                )
                countif_cache = (
                    _build_countif_cache(df_raw, i, alumnos_asig_idx)
                    if alumnos_asig_idx is not None
                    else {}
                )

                j = i + 1
                while j < n_rows:
                    row  = df_raw.iloc[j]
                    asig = _cell(row, asig_idx)
                    if not asig or asig.lower() in ("nan", "none", "total", "subtotal"):
                        break
                    entry = _read_catalog_entry(row, j, indices, ws_data, countif_cache)
                    rows.append(entry)
                    j += 1

                i = j + 1

        rows.sort(key=_catalog_sort_key)
        logger.debug("[catalog] Catálogo cargado: %d asignaturas", len(rows))

        if _wb_data is not None:
            try:
                _wb_data.close()
            except Exception:
                pass

        return rows

    except Exception as e:
        logger.warning("[catalog] Error leyendo catálogo de asignaturas: %s", e)
        return []
