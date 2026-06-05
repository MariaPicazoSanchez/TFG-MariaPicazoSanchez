"""
Inserción y actualización de estudiantes en Excel.

Exporta:
  - append_user_to_excel          — añade estudiante al Excel
  - handle_save_student_query     — handler Streamlit para guardar alumno
"""


import json
import logging
import os
from collections import Counter
from copy import copy

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from ._insert_helpers import _pick_col, _sheet_exists, first_sheet_name
from ._insert_row_builders import build_existing_sheet_row, build_new_sheet_row

logger = logging.getLogger("movilidad_persistence")


# ─────────────────────────────────────────────────────────────────────────────
# Erasmus IN: inserción por asignatura
# ─────────────────────────────────────────────────────────────────────────────

def _append_erasmus_in_with_subjects(
    xlsx_path: str, row_data: dict, target_sheet: str, lat, lon
) -> tuple[bool, str | None]:
    """
    Añade una fila por asignatura para un estudiante Erasmus IN.
    Devuelve (ok, error_msg).
    """
    materias = row_data.get("materias_in", [])
    if not materias:
        return True, None

    nombre_completo    = f"{row_data.get('nombre', '')} {row_data.get('apellidos', '')}".strip()
    origen             = row_data.get("pais_in", "")
    universidad_origen = row_data.get("destino_origen", "")
    la_default         = row_data.get("la_in", "")

    cols = ["Asignatura", "Estudiante", "Origen", "Universidad Origen", "Cuat", "Firmado", "LA"]

    def _safe_int(v, default=0):
        try:
            return int(v) if v not in (None, "") else default
        except (TypeError, ValueError):
            return default

    rows_to_add = [
        {
            "Asignatura":        m.get("asignatura", "").strip(),
            "Estudiante":        nombre_completo,
            "Origen":            origen,
            "Universidad Origen": universidad_origen,
            "Cuat":              m.get("cuat", ""),
            "Firmado":           m.get("firmado", ""),
            "LA":                m.get("link_la", la_default),
            "Cupo":              _safe_int(m.get("cupo"), 0),
        }
        for m in materias
        if m.get("asignatura", "").strip()
    ]

    if not rows_to_add:
        return True, None

    if not _sheet_exists(xlsx_path, target_sheet):
        # Intentar clonar la hoja de curso académico más reciente como plantilla,
        # insertar materias y actualizar el catálogo en la misma sesión openpyxl.
        try:
            from ._erasmus_in_catalog import (
                append_to_catalog,
                clone_sheet_as_new_course,
                extend_tables_ref_to_row,
                find_catalog_in_ws,
                find_materias_header_in_ws,
                insert_materias_rows,
                pick_template_sheet,
            )

            wb_tpl = load_workbook(xlsx_path)
            cloned = False
            try:
                template_name = pick_template_sheet(wb_tpl, exclude=target_sheet)
                if template_name:
                    clone_sheet_as_new_course(wb_tpl, template_name, target_sheet)
                    new_ws = wb_tpl[target_sheet]

                    header_hit = find_materias_header_in_ws(new_ws)
                    if header_hit:
                        header_row, cols_map = header_hit
                        last_materias = insert_materias_rows(
                            new_ws, header_row, cols_map, rows_to_add,
                        )
                        last_catalog = 0
                        try:
                            cat_info = find_catalog_in_ws(new_ws)
                            if cat_info:
                                counts = Counter(
                                    f["Asignatura"] for f in rows_to_add if f["Asignatura"]
                                )
                                # Asignaturas del estudiante: matr=count, cupo=valor del form
                                entries = [
                                    {
                                        "asignatura": f["Asignatura"],
                                        "cuat":       f["Cuat"],
                                        "matriculados": counts[f["Asignatura"]],
                                        "cupo":       f.get("Cupo", 0),
                                        "_from_student": True,
                                    }
                                    for f in rows_to_add if f["Asignatura"]
                                ]
                                append_to_catalog(new_ws, cat_info, entries)
                                last_catalog = cat_info["data_end"]
                        except Exception as e:
                            logger.warning("No se pudo actualizar el catálogo: %s", e)

                        extend_tables_ref_to_row(new_ws, max(last_materias, last_catalog))

                        wb_tpl.save(xlsx_path)
                        cloned = True
            finally:
                wb_tpl.close()

            if cloned:
                return True, None
        except PermissionError:
            return False, "El archivo está abierto en otra aplicación."
        except Exception as e:
            logger.warning("No se pudo clonar plantilla para '%s': %s", target_sheet, e)

        if not _sheet_exists(xlsx_path, target_sheet):
            df_new = pd.DataFrame(rows_to_add, columns=cols)
            mode = "a" if os.path.exists(xlsx_path) else "w"
            try:
                with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode=mode) as w:
                    df_new.to_excel(w, sheet_name=target_sheet, index=False)
            except PermissionError:
                return False, "El archivo está abierto en otra aplicación."
            except Exception as e:
                return False, f"Error creando la hoja '{target_sheet}': {e}"
            return True, None

    try:
        from persistence.excel_update import _find_table_in_workbook, MATERIAS_HEADER_ALIASES, MATERIAS_REQUIRED

        table_info = _find_table_in_workbook(
            xlsx_path, MATERIAS_HEADER_ALIASES, MATERIAS_REQUIRED,
            extra_min_matches=2,
            extras_pool={"origen", "universidad_origen", "cuat", "firmado"},
            target_sheet=target_sheet,
        )

        if not table_info:
            df_existing = pd.read_excel(xlsx_path, sheet_name=target_sheet, engine="openpyxl")
            df_existing.columns = [str(c).strip() for c in df_existing.columns]
            df_new = pd.DataFrame(rows_to_add, columns=cols)
            for col in cols:
                if col not in df_existing.columns:
                    df_existing[col] = None
            for col in df_existing.columns:
                if col not in df_new.columns:
                    df_new[col] = None
            cols_order = list(df_existing.columns)
            df_out = pd.concat([df_existing, df_new[cols_order]], ignore_index=True)
            with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
                df_out.to_excel(w, sheet_name=target_sheet, index=False)
            return True, None

        wb = load_workbook(xlsx_path)
        ws = wb[target_sheet]

        c_asig = table_info.cols.get("asignatura")
        c_est  = table_info.cols.get("estudiante")
        c_ori  = table_info.cols.get("origen")
        c_uni  = table_info.cols.get("universidad_origen")
        c_cuat = table_info.cols.get("cuat")
        c_fir  = table_info.cols.get("firmado")
        c_la   = table_info.cols.get("link_la")

        last_row = table_info.data_end
        if c_asig:
            for r in range(table_info.data_start, ws.max_row + 1):
                v = ws.cell(row=r, column=c_asig).value
                if v is not None and str(v).strip() != "" and r <= table_info.data_end + 50:
                    last_row = max(last_row, r)

        insert_row = last_row + 1
        fmt_row = last_row if last_row >= table_info.data_start else table_info.header_row

        # Restringimos la copia de estilos al rango de columnas de la propia
        # tabla de materias. Si copiamos para todas las columnas de la hoja,
        # extendemos el formato/banding de la tabla lateral de asignaturas
        # (catálogo: Asignatura/Cuat/Matriculados/Cupo) a las filas nuevas
        # del alumno, que no deberían tocarse.
        _mat_cols = [v for v in table_info.cols.values() if isinstance(v, int) and v > 0]
        if _mat_cols:
            col_min, col_max = min(_mat_cols), max(_mat_cols)
        else:
            col_min, col_max = 1, ws.max_column

        for i, fila in enumerate(rows_to_add):
            r = insert_row + i
            for col_idx in range(col_min, col_max + 1):
                src = ws.cell(row=fmt_row, column=col_idx)
                dst = ws.cell(row=r, column=col_idx)
                if src.has_style:
                    dst._style = copy(src._style)
            if c_asig: ws.cell(row=r, column=c_asig).value = fila["Asignatura"]
            if c_est:  ws.cell(row=r, column=c_est).value  = fila["Estudiante"]
            if c_ori:  ws.cell(row=r, column=c_ori).value  = fila["Origen"]
            if c_uni:  ws.cell(row=r, column=c_uni).value  = fila["Universidad Origen"]
            if c_cuat: ws.cell(row=r, column=c_cuat).value = fila["Cuat"]
            if c_fir:  ws.cell(row=r, column=c_fir).value  = fila["Firmado"]
            if c_la:   ws.cell(row=r, column=c_la).value   = fila["LA"]

        last_inserted = insert_row + len(rows_to_add) - 1
        from openpyxl.utils import range_boundaries, get_column_letter
        for tbl in ws.tables.values():
            min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
            # Solo extender la tabla de materias (la que contiene su cabecera
            # en table_info.header_row). No tocar otras tablas de la hoja
            # (p.ej. el catálogo lateral de asignaturas), para no destruir su
            # formato ni recortar/extender su rango incorrectamente.
            if min_row != table_info.header_row:
                continue
            new_max = max(last_inserted, max_row)
            tbl.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max}"

        # Actualiza el catálogo lateral: añade asignaturas nuevas con matr = cuenta
        # en las materias del estudiante; añade además las asignaturas vistas en
        # otros cursos (matr=0, cupo=0) si no estaban ya en este curso, para que
        # el desplegable de sugerencias muestre opciones cruzadas.
        try:
            from ._erasmus_in_catalog import (
                append_to_catalog,
                find_catalog_in_ws,
            )
            cat_info = find_catalog_in_ws(ws)
            if cat_info:
                counts = Counter(
                    fila["Asignatura"] for fila in rows_to_add if fila["Asignatura"]
                )
                entries = [
                    {
                        "asignatura":   fila["Asignatura"],
                        "cuat":         fila["Cuat"],
                        "matriculados": counts[fila["Asignatura"]],
                        "cupo":         fila.get("Cupo", 0),
                        "_from_student": True,
                    }
                    for fila in rows_to_add if fila["Asignatura"]
                ]
                append_to_catalog(ws, cat_info, entries)
        except Exception as e:
            logger.warning("No se pudo actualizar el catálogo de asignaturas: %s", e)

        wb.save(xlsx_path)
        wb.close()
        return True, None

    except PermissionError:
        return False, "El archivo está abierto en otra aplicación."
    except Exception as e:
        return False, f"Error guardando: {e}"


# ─────────────────────────────────────────────────────────────────────────────
# Inserción principal
# ─────────────────────────────────────────────────────────────────────────────

def append_user_to_excel(
    xlsx_path: str, tipo: str, row_data: dict, sheet_name: str | None
) -> tuple[bool, str | None]:
    """
    Añade una fila al Excel en la hoja `sheet_name`.
    - Erasmus IN con materias: una fila por asignatura.
    - Resto: una fila con la info del estudiante.
    Devuelve (ok, error_msg).
    """
    if not xlsx_path or not os.path.exists(xlsx_path):
        return False, f"No existe el Excel: {xlsx_path}"

    target_sheet = (sheet_name or "").strip() or first_sheet_name(xlsx_path)

    lat = lon = None
    if "coordenadas" in row_data and isinstance(row_data["coordenadas"], (tuple, list)) and len(row_data["coordenadas"]) == 2:
        lat, lon = row_data["coordenadas"]
    else:
        lat = row_data.get("lat")
        lon = row_data.get("lon")

    if tipo == "Erasmus IN" and row_data.get("materias_in"):
        ok, err = _append_erasmus_in_with_subjects(xlsx_path, row_data, target_sheet, lat, lon)
        if ok:
            _ensure_university_in_coords(xlsx_path, tipo, row_data)
        return ok, err

    # ── Hoja NO existe → crear con columnas estándar ──────────────────────────
    if not _sheet_exists(xlsx_path, target_sheet):
        new, need_cols = build_new_sheet_row(tipo, row_data, lat, lon)
        out = pd.DataFrame([new], columns=need_cols)
        mode = "a" if os.path.exists(xlsx_path) else "w"
        try:
            with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode=mode) as w:
                out.to_excel(w, sheet_name=target_sheet, index=False)
        except PermissionError:
            return False, "El archivo está abierto en otra aplicación."
        except Exception as e:
            return False, f"Error creando la hoja '{target_sheet}': {e}"
        return True, None

    # ── Hoja SÍ existe → leer, mapear y añadir ───────────────────────────────
    try:
        df = pd.read_excel(xlsx_path, sheet_name=target_sheet, engine="openpyxl")
    except Exception as e:
        return False, f"Error leyendo hoja '{target_sheet}': {e}"

    df.columns = [str(c).strip() for c in df.columns]
    new_row, cols_order = build_existing_sheet_row(tipo, df, row_data, lat, lon)

    out = pd.concat(
        [df, pd.DataFrame([new_row])], ignore_index=True
    ).reindex(columns=cols_order)

    # SICUE OUT no tiene hoja "Coordenadas" aparte: la columna Coordenadas
    # vive en cada hoja de curso. Si la universidad ya aparece en otro alumno
    # se copia su valor; si no, se geocodifica con sufijo " (auto)".
    if tipo == "SICUE OUT":
        universidad = (row_data.get("destino_origen") or "").strip()
        coords_col = next(
            (c for c in cols_order if str(c).strip().lower() == "coordenadas"),
            None,
        )
        if universidad and coords_col is not None:
            try:
                from ._coords_sheet import resolve_sicue_coords_for_universidad
                coords_val = resolve_sicue_coords_for_universidad(xlsx_path, universidad)
                if coords_val:
                    out.at[len(out) - 1, coords_col] = coords_val
            except Exception as e:
                logger.warning("[coords-sicue] No se pudo resolver coords: %s", e)

    # Erasmus IN / Erasmus OUT: la columna Coordenadas de la hoja del curso se
    # deja vacía a propósito; la fuente de verdad es la hoja "Coordenadas".

    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
            out.to_excel(w, sheet_name=target_sheet, index=False)
    except TypeError:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="w") as w:
            out.to_excel(w, sheet_name=target_sheet, index=False)
    except PermissionError:
        return False, "El archivo está abierto en otra aplicación."
    except Exception as e:
        return False, f"Error guardando: {e}"

    # Si la universidad es nueva, geocodificarla y añadirla a la hoja
    # "Coordenadas" con el sufijo " (auto)" para que el coordinador sepa que
    # son coordenadas auto-generadas y pueda revisarlas.
    _ensure_university_in_coords(xlsx_path, tipo, row_data)

    return True, None


def _ensure_university_in_coords(xlsx_path: str, tipo: str, row_data: dict) -> None:
    """Wrapper tolerante a fallos para no bloquear el guardado del alumno."""
    try:
        from ._coords_sheet import ensure_university_in_coords_sheet
        universidad = (row_data.get("destino_origen") or "").strip()
        if tipo == "Erasmus IN":
            pais = (row_data.get("pais_in") or "").strip()
        elif tipo == "Erasmus OUT":
            pais = (row_data.get("pais_out") or "").strip()
        else:
            pais = ""
        if universidad:
            ensure_university_in_coords_sheet(xlsx_path, tipo, universidad, pais)
    except Exception as e:
        logger.warning("[coords-sheet] No se pudo asegurar universidad: %s", e)


# ─────────────────────────────────────────────────────────────────────────────
# Exportación de materias
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# Handler Streamlit: guardar alumno desde query params
# ─────────────────────────────────────────────────────────────────────────────

def handle_save_student_query() -> None:
    from ui.popup_helpers import _normalize_estudiantes

    params = st.query_params
    if "save_student" not in params:
        return

    def _qp_val(p, key):
        v = p.get(key)
        if v is None:
            return None
        if isinstance(v, list):
            return v[0] if v else ""
        return str(v)

    programa = _qp_val(params, "programa")
    row_id   = _qp_val(params, "row_id")
    idx_str  = _qp_val(params, "idx")

    if programa is None or row_id is None or idx_str is None:
        st.error("Faltan parámetros para guardar el alumno.")
        return

    try:
        idx = int(idx_str)
    except ValueError:
        st.error("Índice de estudiante no válido.")
        return

    campos: dict = {}
    for key in ("estudiante", "email", "curso", "cuatrimestre",
                "duracion_meses", "gestion_LA", "coordinador_destino",
                "link_la", "link_plan"):
        v = _qp_val(params, key)
        if v is not None:
            campos[key] = v

    materias_list: list = []
    materias_raw = _qp_val(params, "materias_raw") or ""
    if materias_raw.strip():
        for line in materias_raw.splitlines():
            line = line.strip()
            if not line:
                continue
            if "|" in line:
                asig, cuat = [p.strip() for p in line.split("|", 1)]
            else:
                asig, cuat = line, ""
            if asig:
                materias_list.append({"asignatura": asig, "cuat": cuat})

    config = st.session_state.get("config", {})
    ruta = config.get(programa)
    if ruta is None:
        key_norm = (programa or "").strip().lower()
        norm_map = {(k.strip().lower() if isinstance(k, str) else k): v for k, v in config.items()}
        ruta = norm_map.get(key_norm)
        if ruta is None:
            for k, v in norm_map.items():
                if key_norm and key_norm in str(k):
                    ruta = v
                    break

    if not ruta:
        st.error(f"No se ha encontrado ruta para {programa!r} en config.json")
        return

    try:
        df = pd.read_excel(ruta)
    except Exception as e:
        st.error(f"No se ha podido leer el Excel {ruta}: {e}")
        return

    # Sin columna 'id' → buscar por email/nombre o añadir
    if "id" not in df.columns:
        c_email     = _pick_col(df, "email", "Email")
        c_nombre    = _pick_col(df, "nombre", "Nombre")
        c_apellidos = _pick_col(df, "apellidos", "Apellidos", "apellido1", "apellido2")

        found_idx = None
        email_val = (campos.get("email") or "").strip()
        if c_email and email_val:
            mask = df[c_email].astype(str).str.strip().eq(email_val)
            if mask.any():
                found_idx = df.index[mask][0]

        if found_idx is None:
            nombre_val = (campos.get("estudiante") or "").strip()
            apes_val   = (campos.get("apellidos") or "").strip()
            if c_nombre and nombre_val:
                mask_n = df[c_nombre].astype(str).str.strip().str.lower().eq(nombre_val.lower())
                if c_apellidos and apes_val:
                    mask = mask_n & df[c_apellidos].astype(str).str.strip().str.lower().eq(apes_val.lower())
                else:
                    mask = mask_n
                if mask.any():
                    found_idx = df.index[mask][0]

        if found_idx is not None:
            for k, v in campos.items():
                col = _pick_col(df, k, k.capitalize())
                if col:
                    df.at[found_idx, col] = v
            if materias_list:
                col_m = _pick_col(df, "materias_in", "Materias", "materias")
                if col_m:
                    df.at[found_idx, col_m] = str(materias_list)
            try:
                df.to_excel(ruta, index=False)
                st.session_state["_student_saved"] = True
                st.success("✅ Alumno actualizado correctamente.")
            except Exception as e:
                st.error(f"Error guardando Excel: {e}")
            return

        st.warning("No se encontró 'id' ni coincidencia por email/nombre: se añadirá una nueva fila.")
        try:
            append_user_to_excel(ruta, programa, {**campos, "materias": materias_list}, None)
            st.session_state["_student_saved"] = True
            st.success("✅ Alumno añadido correctamente.")
        except Exception as e:
            st.error(f"Error añadiendo fila: {e}")
        return

    # Con columna 'id' → actualizar lista 'estudiantes'
    mask = df["id"].astype(str) == str(row_id)
    if not mask.any():
        st.error(f"No se ha encontrado la fila con id={row_id} en {programa}.")
        return

    fila_idx = df[mask].index[0]
    est_raw  = df.at[fila_idx, "estudiantes"]

    if est_raw is None or (isinstance(est_raw, str) and not est_raw.strip()):
        lista_est = []
    else:
        try:
            lista_est = _normalize_estudiantes(est_raw)
        except Exception:
            lista_est = []

    if not (0 <= idx < len(lista_est)):
        st.error(f"Índice de estudiante {idx} fuera de rango.")
        return

    est = lista_est[idx] if isinstance(lista_est[idx], dict) else {}
    est.update(campos)
    if programa == "Erasmus IN":
        est["materias_in"] = materias_list

    lista_est[idx] = est
    df.at[fila_idx, "estudiantes"] = json.dumps(lista_est, ensure_ascii=False)

    try:
        df.to_excel(ruta, index=False)
        st.session_state["_student_saved"] = True
        st.success("✅ Alumno actualizado correctamente.")
    except Exception as e:
        st.error(f"No se ha podido guardar en {ruta}: {e}")
