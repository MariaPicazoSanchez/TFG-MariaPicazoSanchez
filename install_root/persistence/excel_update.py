
import logging
import os
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from ._excel_tables import (
    MATERIAS_HEADER_ALIASES,
    MATERIAS_REQUIRED,
    STUDENTS_HEADER_ALIASES,
    STUDENTS_REQUIRED,
    TableInfo,
    _build_header_maps_from_ws,
    _find_col_in_ws_by_aliases,
    _find_table_in_workbook,
    _iter_all_tables_in_workbook,
    normalize_str,
    _norm_header,
)
from ._excel_cells import (
    FIELD_ALIASES,
    _is_invalid_student_name_cell,
    _name_to_scalar,
    _normalize_firmado,
    _recalculate_coords,
    _set_ws_cell_if_field_exists,
    _split_full_name,
    _students_table_is_dynamic_unique,
)

logger = logging.getLogger("movilidad_persistence")

# Re-exports for backward compatibility (data_insert.py imports these from here)
__all__ = [
    "MATERIAS_HEADER_ALIASES",
    "MATERIAS_REQUIRED",
    "_find_table_in_workbook",
    "_recalculate_coords",
    "get_student_name_by_row",
    "util_get_unique_names_from_column",
    "update_student_in_excel",
    "actualizar_excel_materias_para_estudiante",
]


# ─────────────────────────────────────────────────────────────────────────────
# Utilidades de lectura de nombres
# ─────────────────────────────────────────────────────────────────────────────

def get_student_name_by_row(excel_path: str, sheet_name: str, row: int) -> str:
    """Devuelve el nombre real del alumno en la columna B, ignorando fórmulas y errores."""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    v = ws.cell(row=row, column=2).value
    wb.close()
    if v is None:
        return ""
    v = str(v).strip()
    if not v or v.upper() in {"#N/D", "#VALUE!", "#¡DESBORDAMIENTO!", "#SPILL!", "#REF!", "#NAME?"}:
        return ""
    return v


def util_get_unique_names_from_column(
    excel_path: str, sheet_name: str, col_letter: str, row_start: int = 2, row_end: int = 479
) -> list[str]:
    """Lista de nombres únicos de la columna col_letter, ignorando vacíos y errores."""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    nombres: list[str] = []
    vistos: set[str] = set()
    for fila in range(row_start, row_end + 1):
        v = ws[f"{col_letter}{fila}"].value
        if v is None:
            continue
        v = str(v).strip()
        if not v or v.upper() in {"#N/D", "#VALUE!", "#¡DESBORDAMIENTO!", "#SPILL!", "#REF!", "#NAME?"}:
            continue
        if v not in vistos:
            vistos.add(v)
            nombres.append(v)
    wb.close()
    return nombres


# ─────────────────────────────────────────────────────────────────────────────
# Búsqueda de alumno en hoja
# ─────────────────────────────────────────────────────────────────────────────

def _search_student_in_ws(ws, table_info: TableInfo, email_target: str, cand_norm: str) -> Optional[int]:
    """Devuelve la fila 1-based donde está el alumno en ws, o None.

    Busca primero por nombre (identificador único por alumno) y solo si no
    encuentra por nombre, cae en la búsqueda por email. Esto evita que el
    email del coordinador (compartido entre varios alumnos en Erasmus OUT)
    devuelva el primer alumno de la lista en lugar del alumno correcto.
    """
    email_col  = table_info.cols.get("email")
    nombre_col = table_info.cols.get("nombre")
    ap1_col    = table_info.cols.get("apellido1")
    ap2_col    = table_info.cols.get("apellido2")
    search_end = ws.max_row

    # 1) Búsqueda por nombre (tiene prioridad porque es único por alumno)
    if nombre_col and cand_norm:
        for r in range(table_info.data_start, search_end + 1):
            cell_name = _name_to_scalar(ws.cell(row=r, column=nombre_col).value)
            if _is_invalid_student_name_cell(cell_name):
                continue
            if ap1_col:
                ap1_val = _name_to_scalar(ws.cell(row=r, column=ap1_col).value)
                ap2_val = _name_to_scalar(ws.cell(row=r, column=ap2_col).value) if ap2_col else ""
                parts = [p for p in [cell_name, ap1_val, ap2_val] if p and not _is_invalid_student_name_cell(p)]
                name_norm = normalize_str(" ".join(parts))
            else:
                name_norm = normalize_str(cell_name)
            if name_norm == cand_norm:
                return r

    # 2) Fallback: búsqueda por email (solo si no se encontró por nombre)
    if email_col and email_target:
        for r in range(table_info.data_start, search_end + 1):
            cell_email = _name_to_scalar(ws.cell(row=r, column=email_col).value)
            if not _is_invalid_student_name_cell(cell_email) and str(cell_email).strip().lower() == email_target:
                return r

    return None


# ─────────────────────────────────────────────────────────────────────────────
# Actualización de alumno
# ─────────────────────────────────────────────────────────────────────────────

def update_student_in_excel(
    excel_path: str,
    row_index: str,
    idx: int,
    data: dict,
    old_email: str = None,
    old_nombre: str = None,
    target_sheet: str = "",
) -> bool:
    """
    Actualiza SOLO la fila del alumno dentro de la tabla de alumnos.
    Si target_sheet está definida, busca únicamente en esa hoja.
    """
    logger.debug("[update_student_in_excel] excel_path=%s row_index=%s idx=%s", excel_path, row_index, idx)

    email_target = (data.get("old_email") or old_email or "").strip().lower()
    full_name_raw = (data.get("old_nombre") or old_nombre or data.get("estudiante") or "").strip()
    cand_norm = normalize_str(_name_to_scalar(full_name_raw))

    wb = None
    try:
        wb = load_workbook(excel_path)

        table_info_found = None
        row_found = None
        for ti in _iter_all_tables_in_workbook(
            wb, STUDENTS_HEADER_ALIASES, STUDENTS_REQUIRED,
            extra_min_matches=1,
            extras_pool={"cuatrimestre", "universidad_origen", "coordenadas", "email"},
        ):
            if target_sheet and ti.sheet_name != target_sheet:
                logger.debug("[update_student_in_excel] Saltando hoja=%r (buscando en %r)", ti.sheet_name, target_sheet)
                continue
            logger.debug("[update_student_in_excel] Probando tabla hoja=%r header=%d data=%d..%d cols=%s",
                         ti.sheet_name, ti.header_row, ti.data_start, ti.data_end, list(ti.cols.keys()))
            ws_ti = wb[ti.sheet_name]
            r = _search_student_in_ws(ws_ti, ti, email_target, cand_norm)
            if r is not None:
                table_info_found = ti
                row_found = r
                logger.debug("[update_student_in_excel] Alumno encontrado en hoja=%r fila=%d", ti.sheet_name, r)
                break

        if table_info_found is None or row_found is None:
            logger.warning(
                "[update_student_in_excel] No se encontró ninguna fila válida "
                "(email_target=%r, cand_norm=%r).", email_target, cand_norm
            )
            return False

        ws = wb[table_info_found.sheet_name]
        norm_to_col_1based, raw_headers = _build_header_maps_from_ws(ws, table_info_found.header_row - 1)
        logger.debug("[update_student_in_excel] headers detectados: %s", raw_headers)

        is_dynamic_matrix = _students_table_is_dynamic_unique(ws, table_info_found)

        # Para SICUE OUT capturamos el destino actual antes de aplicar los
        # cambios: la columna Coordenadas de la hoja se recalcula solo si la
        # universidad de destino cambia (ver más abajo).
        _col_destino_pre = _find_col_in_ws_by_aliases(
            norm_to_col_1based, ["destino", "Destino"],
        )
        _old_destino = ""
        if _col_destino_pre:
            _v = ws.cell(row=row_found, column=_col_destino_pre).value
            _old_destino = "" if _v is None else str(_v).strip()

        for field in FIELD_ALIASES.keys():
            if is_dynamic_matrix and field in ("estudiante", "nombre"):
                continue
            _set_ws_cell_if_field_exists(ws, row_found, norm_to_col_1based, field, data)

        if not is_dynamic_matrix:
            full = (data.get("estudiante") or "").strip()
            if full:
                nombre, ape1, ape2 = _split_full_name(full)
                col_nombre = _find_col_in_ws_by_aliases(norm_to_col_1based, ["nombre", "estudiante", "nombre completo"])
                col_ape1   = _find_col_in_ws_by_aliases(norm_to_col_1based, ["apellido1", "apellido_1"])
                col_ape2   = _find_col_in_ws_by_aliases(norm_to_col_1based, ["apellido2", "apellido_2"])

                if col_nombre:
                    header_norm = _norm_header(ws.cell(row=table_info_found.header_row, column=col_nombre).value)
                    if header_norm in ("nombre completo", "estudiante"):
                        ws.cell(row=row_found, column=col_nombre).value = full
                    elif header_norm == "nombre" and col_ape1:
                        ws.cell(row=row_found, column=col_nombre).value = nombre
                    else:
                        ws.cell(row=row_found, column=col_nombre).value = full

                if col_ape1 and ape1:
                    ws.cell(row=row_found, column=col_ape1).value = ape1
                if col_ape2 and ape2:
                    ws.cell(row=row_found, column=col_ape2).value = ape2

        # Inferir programa por el nombre del fichero (no hay campo explícito).
        programa_guess = ""
        low = (excel_path or "").lower()
        if "erasmus out" in low:
            programa_guess = "Erasmus OUT"
        elif "erasmus in" in low:
            programa_guess = "Erasmus IN"
        elif "sicue" in low:
            programa_guess = "SICUE OUT"

        # SICUE OUT: la columna Coordenadas vive en la hoja del curso. Si la
        # universidad de destino ha cambiado, recalcular el valor copiándolo de
        # otra fila con la misma universidad o, en su defecto, geocodificando.
        if programa_guess == "SICUE OUT":
            new_destino = (data.get("destino") or "").strip()
            col_coords = _find_col_in_ws_by_aliases(
                norm_to_col_1based, ["coordenadas", "Coordenadas"],
            )
            if (col_coords and new_destino
                    and new_destino.lower() != _old_destino.lower()):
                try:
                    from ._coords_sheet import resolve_sicue_coords_for_universidad
                    coords_val = resolve_sicue_coords_for_universidad(
                        excel_path, new_destino,
                        exclude_sheet=table_info_found.sheet_name,
                        exclude_row=row_found,
                    )
                    if coords_val:
                        ws.cell(row=row_found, column=col_coords).value = coords_val
                except Exception as e:
                    logger.warning("[coords-sicue] No se pudo resolver coords: %s", e)

        wb.save(excel_path)
        logger.info("[update_student_in_excel] Guardado OK (misma fila, in-place).")

        # Erasmus IN / OUT: añadir la universidad a la hoja "Coordenadas" si
        # es nueva (con coordenadas auto-geocodificadas marcadas " (auto)").
        # SICUE OUT no tiene esa hoja, así que el helper saldrá sin hacer nada.
        try:
            from ._coords_sheet import ensure_university_in_coords_sheet
            uni = (
                (data.get("destino") or data.get("universidad_origen")
                 or data.get("origen") or "").strip()
            )
            pais = (data.get("pais") or "").strip()
            if uni:
                ensure_university_in_coords_sheet(excel_path, programa_guess, uni, pais)
        except Exception as e:
            logger.warning("[coords-sheet] No se pudo asegurar universidad: %s", e)

        return True

    except Exception:
        logger.exception("[update_student_in_excel] Error guardando in-place")
        return False
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 'Coordenadas': escritura del Plan de estudios por universidad
# ─────────────────────────────────────────────────────────────────────────────

# Col E (1-based = 5) reservada para el plan de estudios, en paralelo a col D
# (responsable) que ya se escribe desde ui/new_user/_helpers.py.
PLAN_ESTUDIOS_COL_XLSX = 5


def _detect_uni_col_coordenadas(path: str, default_col_uni_0based: int) -> int:
    """Detecta en qué columna (0-based) está la universidad en la hoja 'Coordenadas'.

    Si la primera fila tiene una cabecera reconocible ('Universidad', 'Universidade',
    'University'), usa esa. Si no, devuelve el default proporcionado.
    """
    try:
        df = pd.read_excel(path, sheet_name="Coordenadas", header=None, dtype=str)
        if len(df) == 0:
            return default_col_uni_0based
        first_row = [str(v).strip().lower() for v in df.iloc[0]]
        _words_uni = {
            "universidad", "universidade", "university",
            "universidad destino", "universidad origen",
        }
        for i, v in enumerate(first_row):
            if v in _words_uni:
                return i
    except Exception:
        pass
    return default_col_uni_0based


def write_plan_estudios_for_university(
    path: str,
    universidad: str,
    plan_estudios: str,
    default_col_uni_0based: int = 1,
) -> bool:
    """
    Actualiza el plan de estudios de una universidad en la hoja 'Coordenadas'
    del fichero Excel indicado. Escribe en col E (1-based = 5).

    default_col_uni_0based: índice 0-based de la columna Universidad.
        Erasmus OUT → 0
        Erasmus IN  → 1

    Returns True si se encontró la universidad y se escribió, False si no.
    """
    if not path or not universidad:
        return False
    if not os.path.exists(path):
        return False

    col_uni_0 = _detect_uni_col_coordenadas(path, default_col_uni_0based)
    col_uni_1 = col_uni_0 + 1  # openpyxl es 1-based

    wb = None
    try:
        wb = load_workbook(path)
        if "Coordenadas" not in wb.sheetnames:
            return False
        ws = wb["Coordenadas"]
        target  = (universidad or "").strip().lower()
        value   = (plan_estudios or "").strip() or None
        written = False
        # Actualizar TODAS las filas de esa universidad (puede aparecer varias
        # veces en Coordenadas, una por estudiante/historial).
        for r_idx in range(1, ws.max_row + 1):
            uni_val = str(ws.cell(row=r_idx, column=col_uni_1).value or "").strip()
            if uni_val and uni_val.lower() == target:
                ws.cell(row=r_idx, column=PLAN_ESTUDIOS_COL_XLSX).value = value
                written = True
        if written:
            wb.save(path)
        return written
    except Exception:
        logger.exception("write_plan_estudios_for_university: error para %r", universidad)
        return False
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# Helpers internos de materias
# ─────────────────────────────────────────────────────────────────────────────

def _cuat_to_cell_value(s: str):
    """Convierte el valor del cuatrimestre al tipo nativo más apropiado.

    '1' → 1 (int), '1.0' → 1 (int), '2.0' → 2 (int), '1.5' → 1.5 (float),
    'A' → 'A' (str).

    Evita que el cuatrimestre se escriba como texto '1.0' en lugar del número
    entero 1, lo que cambia el formato de la celda en Excel y puede romper la
    detección de columnas al recargar.
    """
    if not s:
        return s
    try:
        f = float(str(s).replace(",", "."))
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return str(s).strip()


# ─────────────────────────────────────────────────────────────────────────────
# Actualización de materias
# ─────────────────────────────────────────────────────────────────────────────

def actualizar_excel_materias_para_estudiante(
    materias_in, est: dict, materias_path: str, sheet_name: str = ""
) -> None:
    """
    Edita en la misma fila las materias existentes del alumno y añade nuevas
    debajo del final de la tabla.
    """
    if not materias_path:
        raise ValueError("No se ha recibido la ruta del Excel de materias.")
    if not os.path.exists(materias_path):
        raise FileNotFoundError(f"El archivo de materias no existe: {materias_path}")

    nombre_nuevo   = str(est.get("estudiante") or "").strip()
    nombre_antiguo = str(est.get("old_nombre") or est.get("old_estudiante") or "").strip()
    origen_default = str(est.get("pais") or est.get("origen") or "").strip()
    uni_default    = str(
        est.get("universidad_origen") or est.get("destino") or est.get("Centro")
        or est.get("centro") or est.get("universidad") or ""
    ).strip()
    cuat_default    = _cuat_to_cell_value(str(est.get("cuat") or "").strip())
    firmado_default = _normalize_firmado(est.get("firmado", ""))
    la_default      = str(est.get("link_la") or "").strip()

    if not nombre_nuevo and not nombre_antiguo:
        raise ValueError("No hay nombre de estudiante para actualizar materias.")

    table_info = _find_table_in_workbook(
        materias_path, MATERIAS_HEADER_ALIASES, MATERIAS_REQUIRED,
        extra_min_matches=2,
        extras_pool={"origen", "universidad_origen", "cuat", "firmado"},
        target_sheet=sheet_name or "",
    )
    if not table_info:
        raise RuntimeError("No se ha encontrado tabla de Materias IN en el Excel.")

    wb = load_workbook(materias_path)
    try:
        ws = wb[table_info.sheet_name]

        c_asig = table_info.cols.get("asignatura")
        c_est  = table_info.cols.get("estudiante")
        c_ori  = table_info.cols.get("origen")
        c_cuat = table_info.cols.get("cuat")
        c_fir  = table_info.cols.get("firmado")
        c_la   = table_info.cols.get("link_la")

        # Detección robusta de columna universidad_origen
        norm_to_col, raw_headers = _build_header_maps_from_ws(ws, table_info.header_row - 1)
        uni_aliases = list(MATERIAS_HEADER_ALIASES.get("universidad_origen", set()))
        uni_aliases += [
            "universidad_origen", "centro", "universidad de origen", "universidad origen",
            "Universidad Origen", "Universidad  Origen", "universidad  origen",
        ]
        uni_aliases_norm = list({normalize_str(a) for a in uni_aliases})
        c_uni = None
        for alias in uni_aliases_norm:
            if alias in norm_to_col:
                c_uni = norm_to_col[alias]
                break

        if not c_uni:
            logger.warning(
                "[materias] No se detectó columna de universidad de origen. "
                "Alias buscados=%s | Cabeceras detectadas=%s", uni_aliases_norm, raw_headers
            )
        else:
            logger.debug("[materias] Columna universidad detectada: c_uni=%s header=%s", c_uni, raw_headers.get(c_uni))

        # Normalizar payload
        nuevas: list[dict] = []
        for m in (materias_in or []):
            if not isinstance(m, dict):
                continue
            asig = str(m.get("asignatura") or m.get("nombre") or "").strip()
            if asig and (asig.startswith("[") or asig.startswith("{")):
                try:
                    import json as _j
                    _inner = _j.loads(asig)
                    if isinstance(_inner, list) and _inner and isinstance(_inner[0], dict):
                        asig = str(_inner[0].get("asignatura") or _inner[0].get("nombre") or asig).strip()
                    elif isinstance(_inner, dict):
                        asig = str(_inner.get("asignatura") or _inner.get("nombre") or asig).strip()
                except Exception:
                    pass
            if not asig:
                continue
            item_origen = str(m.get("origen") or "").strip()
            item_uni    = str(m.get("universidad_origen") or m.get("centro") or "").strip()
            nuevas.append({
                "asignatura":        asig,
                "estudiante":        nombre_nuevo or nombre_antiguo,
                "origen":            item_origen if item_origen else origen_default,
                "universidad_origen": item_uni if item_uni else uni_default,
                "cuat":              str(m.get("cuat") or "").strip(),
                "firmado":           _normalize_firmado(m.get("firmado", "")),
                "link_la":           str(m.get("link_la") or m.get("la") or "").strip(),
            })
        logger.debug("[materias] nuevas=%d asignaturas a guardar", len(nuevas))

        if not nuevas:
            logger.debug("[materias] No hay materias válidas; no se modifica Excel.")
            return

        # Buscar filas existentes del alumno
        rows_student: list[int] = []
        lookup_names = [x for x in [nombre_antiguo, nombre_nuevo] if x]
        if c_est:
            for candidate in lookup_names:
                cand_norm = normalize_str(candidate)
                rows = [
                    r for r in range(table_info.data_start, ws.max_row + 1)
                    if normalize_str(ws.cell(row=r, column=c_est).value) == cand_norm
                ]
                if rows:
                    rows_student = rows
                    logger.debug("[materias] Filas existentes para '%s': %s", candidate, rows_student)
                    break

        # Editar filas existentes (misma fila)
        common = min(len(rows_student), len(nuevas))
        for i in range(common):
            r    = rows_student[i]
            fila = nuevas[i]
            old_ori = ws.cell(row=r, column=c_ori).value if c_ori else None
            old_uni = ws.cell(row=r, column=c_uni).value if c_uni else None

            if c_asig:
                ws.cell(row=r, column=c_asig).value = fila["asignatura"]
            if c_est:
                ws.cell(row=r, column=c_est).value = fila["estudiante"]
            if c_cuat and cuat_default:
                ws.cell(row=r, column=c_cuat).value = cuat_default  # ya convertido a int/float por _cuat_to_cell_value
            if c_fir:
                ws.cell(row=r, column=c_fir).value = firmado_default
            if c_la and la_default:
                ws.cell(row=r, column=c_la).value = la_default

            if c_ori:
                v_ori = fila["origen"].strip() if isinstance(fila["origen"], str) else fila["origen"]
                ws.cell(row=r, column=c_ori).value = v_ori if v_ori else old_ori

            if c_uni:
                v_uni = fila["universidad_origen"].strip() if isinstance(fila["universidad_origen"], str) else fila["universidad_origen"]
                ws.cell(row=r, column=c_uni).value = v_uni if v_uni else old_uni

            logger.debug("[materias] Editada fila %d: %s", r, fila["asignatura"])

        # Eliminar filas sobrantes (compactar sin insert/delete_rows)
        if len(rows_student) > len(nuevas):
            sobrantes     = rows_student[len(nuevas):]
            cols_to_clear = [c for c in [c_asig, c_est, c_ori, c_uni, c_cuat, c_fir, c_la] if c]
            num_sobrantes = len(sobrantes)
            first_sobrante = sobrantes[0]

            # Limitar el escaneo al rango real de la tabla de materias.
            # Usar ws.max_row como límite puede alcanzar otras tablas que
            # estén en la misma hoja (p.ej. la tabla de alumnos), lo que
            # provocaría que el compactado sobreescriba/desplace filas de
            # datos de alumnos y corrompa coordenadas y cuatrimestre.
            _compact_limit = max(table_info.data_end, sobrantes[-1])
            last_data_row = table_info.data_start - 1
            for r in range(table_info.data_start, _compact_limit + 1):
                if any(ws.cell(row=r, column=col).value not in (None, "") for col in cols_to_clear):
                    last_data_row = r

            if last_data_row >= first_sobrante + num_sobrantes:
                for r in range(first_sobrante, last_data_row - num_sobrantes + 1):
                    source = r + num_sobrantes
                    for col in cols_to_clear:
                        ws.cell(row=r, column=col).value = ws.cell(row=source, column=col).value
                    logger.debug("[materias] Subida fila %d -> %d", source, r)
                for r in range(last_data_row - num_sobrantes + 1, last_data_row + 1):
                    for col in cols_to_clear:
                        ws.cell(row=r, column=col).value = None
                    logger.debug("[materias] Limpiada cola fila %d", r)
            else:
                for r in sobrantes:
                    for col in cols_to_clear:
                        ws.cell(row=r, column=col).value = None
                    logger.debug("[materias] Vaciada fila sobrante %d", r)

        # Añadir filas nuevas
        if len(nuevas) > len(rows_student):
            pendientes = nuevas[len(rows_student):]
            # Margen de 200 filas más allá del fin detectado para acomodar
            # materias que ya se añadieron en sesiones anteriores; pero sin
            # alcanzar otras tablas de la misma hoja.
            _insert_limit = table_info.data_end + 200
            last_row_with_asig = table_info.data_start - 1
            if c_asig:
                for r in range(table_info.data_start, _insert_limit + 1):
                    v = ws.cell(row=r, column=c_asig).value
                    if v is not None and str(v).strip() != "":
                        last_row_with_asig = r
            insert_at = max(last_row_with_asig + 1, table_info.data_start)

            mat_cols = [c for c in [c_asig, c_est, c_ori, c_uni, c_cuat, c_fir, c_la] if c]

            def _mat_cols_empty(row_num: int) -> bool:
                return all(
                    ws.cell(row=row_num, column=col).value in (None, "")
                    for col in mat_cols
                )

            write_at = insert_at
            while write_at <= _insert_limit and not _mat_cols_empty(write_at):
                write_at += 1

            valor_uni_existente = None
            if rows_student and c_uni:
                v = ws.cell(row=rows_student[-1], column=c_uni).value
                if v:
                    valor_uni_existente = str(v).strip()

            for i, fila in enumerate(pendientes):
                r        = write_at + i
                valor_uni = fila.get("universidad_origen") or valor_uni_existente or uni_default
                if c_asig: ws.cell(row=r, column=c_asig).value = fila["asignatura"]
                if c_est:  ws.cell(row=r, column=c_est).value  = fila["estudiante"]
                if c_ori:  ws.cell(row=r, column=c_ori).value  = fila["origen"] or ""
                if c_uni:
                    ws.cell(row=r, column=c_uni).value = valor_uni or ""
                    logger.debug("[materias] Escribiendo universidad en fila %d, col %d: %s", r, c_uni, valor_uni)
                else:
                    logger.warning("[materias] No se encontró columna universidad en fila %d. Valor: %s", r, valor_uni)
                if c_cuat and cuat_default: ws.cell(row=r, column=c_cuat).value = cuat_default
                if c_fir:                   ws.cell(row=r, column=c_fir).value   = firmado_default
                if c_la and la_default:     ws.cell(row=r, column=c_la).value    = la_default
                logger.debug("[materias] Añadida fila %d: %s | universidad_origen=%s", r, fila["asignatura"], valor_uni)

        wb.save(materias_path)
        logger.info("[materias] Guardado OK")

    finally:
        wb.close()
