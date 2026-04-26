"""
Helpers para la hoja Erasmus IN de un curso académico:
  - Detección de la tabla de catálogo (Asignatura / Cuat / Matriculados / Cupo)
  - Clonado de una hoja plantilla (curso anterior) para un curso nuevo
  - Inserción de asignaturas en el catálogo con matriculados/cupo a 0
  - Inserción directa de filas de materias en una hoja recién clonada
"""

from __future__ import annotations

import logging
import re
import unicodedata
from copy import copy, deepcopy

from openpyxl.utils import get_column_letter, range_boundaries

logger = logging.getLogger("movilidad_persistence")


# ─────────────────────────────────────────────────────────────────────────────
# Normalización (idéntica a materias_in_loader._norm)
# ─────────────────────────────────────────────────────────────────────────────

def _norm(s) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s.replace(" ", "")


_ASIG_WORDS = ("asignaturas", "asignatura", "materias", "materia")
_CUAT_WORDS = ("cuat", "cuatrimestre", "periodo", "semestre")
_MATR_WORDS = ("matriculados", "matricula", "nmatric", "inscritos", "matr")
_CUPO_WORDS = ("cupo", "plazas", "capacidad", "cupos", "aforo")


# ─────────────────────────────────────────────────────────────────────────────
# Detección del bloque de catálogo dentro de una hoja openpyxl
# ─────────────────────────────────────────────────────────────────────────────

def find_catalog_in_ws(ws) -> dict | None:
    """
    Localiza la cabecera del catálogo (Asignatura + Matriculados/Cupo) en la hoja.
    Devuelve dict {header_row, asig, cuat, matr, cupo, data_end} con columnas
    1-based (como openpyxl) o None si no hay catálogo.
    """
    max_scan = min(ws.max_row or 0, 200)
    max_col  = ws.max_column or 0
    if not max_scan or not max_col:
        return None

    for r in range(1, max_scan + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        normalized = [_norm(v) for v in row_vals]

        asig_candidates = [i for i, cell in enumerate(normalized) if cell in _ASIG_WORDS]
        if not asig_candidates:
            continue

        best = None
        best_score = -1

        for asig_idx in asig_candidates:
            idxs = {"asig": asig_idx, "cuat": -1, "matr": -1, "cupo": -1}
            for idx in range(asig_idx + 1, min(asig_idx + 10, len(normalized))):
                cell = normalized[idx]
                if cell in _CUAT_WORDS and idxs["cuat"] == -1:
                    idxs["cuat"] = idx
                elif cell in _MATR_WORDS and idxs["matr"] == -1:
                    idxs["matr"] = idx
                elif cell in _CUPO_WORDS and idxs["cupo"] == -1:
                    idxs["cupo"] = idx

            has_matr = idxs["matr"] != -1
            has_cupo = idxs["cupo"] != -1
            score = (has_matr and has_cupo) * 4 + has_matr * 2 + has_cupo * 2
            if score > best_score and (has_matr or has_cupo):
                best_score = score
                best = idxs

        if best is None:
            continue

        data_end = r
        asig_col = best["asig"] + 1
        rr = r + 1
        while rr <= (ws.max_row or 0):
            v = ws.cell(row=rr, column=asig_col).value
            if v is None or str(v).strip() == "" or str(v).strip().lower() in ("nan", "none", "total", "subtotal"):
                break
            data_end = rr
            rr += 1

        return {
            "header_row": r,
            "asig":       asig_col,
            "cuat":       best["cuat"] + 1 if best["cuat"] != -1 else None,
            "matr":       best["matr"] + 1 if best["matr"] != -1 else None,
            "cupo":       best["cupo"] + 1 if best["cupo"] != -1 else None,
            "data_end":   data_end,
        }

    return None


def get_catalog_asignaturas(ws, info: dict) -> set[str]:
    """Devuelve el conjunto de asignaturas ya presentes (normalizadas) en el catálogo."""
    if not info:
        return set()
    asig_col = info["asig"]
    names: set[str] = set()
    for r in range(info["header_row"] + 1, info["data_end"] + 1):
        v = ws.cell(row=r, column=asig_col).value
        k = _norm(v)
        if k:
            names.add(k)
    return names


def append_to_catalog(
    ws,
    info: dict,
    entries: list[dict],
    matr_default: int = 0,
    cupo_default: int = 0,
) -> int:
    """
    Añade asignaturas nuevas al catálogo (las ya presentes se ignoran).
    `entries`: lista de {asignatura, cuat}.
    Devuelve el número de filas insertadas.
    """
    if not info or not entries:
        return 0

    existing = get_catalog_asignaturas(ws, info)
    to_add = []
    seen_norm = set(existing)
    for e in entries:
        name = str(e.get("asignatura", "") or "").strip()
        if not name:
            continue
        key = _norm(name)
        if key in seen_norm:
            continue
        seen_norm.add(key)
        to_add.append({"asignatura": name, "cuat": str(e.get("cuat", "") or "").strip()})

    # Índice rápido para recuperar matriculados/cupo del entry original
    entries_by_key = {}
    for e in entries:
        k = _norm(e.get("asignatura", ""))
        if k and k not in entries_by_key:
            entries_by_key[k] = e

    # Para asignaturas YA presentes en el catálogo que vengan marcadas como
    # "_from_student" actualizamos su cupo (lo que el usuario haya tecleado en
    # el formulario). El resto (sugerencias cruzadas con cupo=0) se respeta.
    if info.get("cupo"):
        for r in range(info["header_row"] + 1, info["data_end"] + 1):
            existing_name = ws.cell(row=r, column=info["asig"]).value
            k = _norm(existing_name)
            entry = entries_by_key.get(k)
            if entry and entry.get("_from_student"):
                cupo_val = entry.get("cupo")
                if cupo_val is not None:
                    ws.cell(row=r, column=info["cupo"]).value = cupo_val

    if not to_add:
        return 0

    # Si el catálogo está vacío (tras clonar) preferimos header_row + 1: su estilo
    # se preservó durante el clonado aunque el valor se borrase. Usar header_row
    # como origen de estilo replicaría el formato de cabecera (negrita + azul) en
    # las filas de datos.
    if info["data_end"] > info["header_row"]:
        fmt_row = info["data_end"]
    else:
        fmt_row = info["header_row"] + 1
    insert_start = info["data_end"] + 1

    # Restringimos la copia de estilo a las columnas del propio catálogo para no
    # pisar el formato de otras tablas contiguas (p. ej. la tabla de materias
    # comparte fila con la cabecera del catálogo).
    cat_cols = [info["asig"]]
    for k in ("cuat", "matr", "cupo"):
        if info.get(k):
            cat_cols.append(info[k])
    col_min, col_max = min(cat_cols), max(cat_cols)

    for i, item in enumerate(to_add):
        r = insert_start + i
        for col_idx in range(col_min, col_max + 1):
            src = ws.cell(row=fmt_row, column=col_idx)
            dst = ws.cell(row=r, column=col_idx)
            if src.has_style:
                dst._style = copy(src._style)
        ws.cell(row=r, column=info["asig"]).value = item["asignatura"]
        if info["cuat"]:
            ws.cell(row=r, column=info["cuat"]).value = item["cuat"]
        orig_entry = entries_by_key.get(_norm(item["asignatura"]), {})
        if info["matr"]:
            ws.cell(row=r, column=info["matr"]).value = orig_entry.get("matriculados", matr_default)
        if info["cupo"]:
            ws.cell(row=r, column=info["cupo"]).value = orig_entry.get("cupo", cupo_default)

    info["data_end"] = insert_start + len(to_add) - 1
    return len(to_add)


def gather_other_course_subjects(wb, current_sheet_name: str) -> list[dict]:
    """
    Recopila las asignaturas presentes en los catálogos de las demás hojas de
    curso académico del workbook (excluyendo `current_sheet_name`). Devuelve
    una lista de {asignatura, cuat} sin duplicados (clave normalizada).

    Sirve para enriquecer el catálogo de un curso nuevo o ya existente con
    asignaturas vistas en otros cursos, de modo que el desplegable de
    sugerencias del editor muestre opciones cruzadas.
    """
    seen: set[str] = set()
    out: list[dict] = []
    for ws in wb.worksheets:
        if ws.title == current_sheet_name:
            continue
        if not _is_academic_year_sheet(ws.title):
            continue
        info = find_catalog_in_ws(ws)
        if not info:
            continue
        asig_col = info["asig"]
        cuat_col = info.get("cuat")
        for r in range(info["header_row"] + 1, info["data_end"] + 1):
            asig = ws.cell(row=r, column=asig_col).value
            name = str(asig).strip() if asig else ""
            if not name:
                continue
            key = _norm(name)
            if not key or key in seen:
                continue
            seen.add(key)
            cuat = ""
            if cuat_col:
                v = ws.cell(row=r, column=cuat_col).value
                cuat = str(v).strip() if v else ""
            out.append({"asignatura": name, "cuat": cuat})
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Clonado de hoja plantilla para un curso académico nuevo
# ─────────────────────────────────────────────────────────────────────────────

_YEAR_RE = re.compile(r"(\d{4})")


def _is_academic_year_sheet(name: str) -> bool:
    return bool(_YEAR_RE.search(str(name)))


def pick_template_sheet(wb, exclude: str | None = None) -> str | None:
    """Devuelve el nombre de la hoja de curso académico más reciente (o None)."""
    candidates = [
        s for s in wb.sheetnames
        if _is_academic_year_sheet(s) and s.lower() != "coordenadas" and s != exclude
    ]
    if not candidates:
        return None

    def _year_key(s: str) -> int:
        m = _YEAR_RE.search(s)
        return int(m.group(1)) if m else 0

    return sorted(candidates, key=_year_key, reverse=True)[0]


def _clear_rows_preserve_style(ws, start_row: int, end_row: int, max_col: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None


def _copy_tables_from_source(wb, src_ws, dst_ws) -> None:
    """
    Copia (deep-copy) los objetos Table de `src_ws` a `dst_ws` renombrándolos
    para que sean únicos dentro del workbook. openpyxl.copy_worksheet no copia
    tablas, por lo que sin esto se perdería el banding/estilo tabular.
    """
    existing_names: set[str] = set()
    for ws in wb.worksheets:
        if ws is dst_ws:
            continue
        existing_names.update(ws.tables.keys())

    for name in list(dst_ws.tables):
        del dst_ws.tables[name]

    # ws.tables.items() devuelve (name, ref_str) en openpyxl 3.x; los objetos
    # Table reales están en .values() o indexando por nombre.
    for src_tbl in list(src_ws.tables.values()):
        src_name = src_tbl.name
        try:
            new_tbl = deepcopy(src_tbl)
        except Exception as e:
            logger.debug("deepcopy falló para Table '%s': %s", src_name, e)
            continue
        unique = src_name
        idx = 2
        while unique in existing_names:
            unique = f"{src_name}_{idx}"
            idx += 1
        new_tbl.name = unique
        new_tbl.displayName = unique
        try:
            dst_ws.add_table(new_tbl)
            existing_names.add(unique)
        except Exception as e:
            logger.debug("No se pudo copiar Table '%s' a hoja clonada: %s", src_name, e)


def clone_sheet_as_new_course(wb, template_name: str, new_name: str) -> None:
    """
    Clona `template_name` con el nombre `new_name` dentro del mismo workbook.
    Vacía las filas de datos de la tabla de materias y del catálogo, conservando
    cabeceras, estilos y objetos Table (cuyo rango se reduce a header + 1 fila).
    """
    from ._excel_tables import MATERIAS_HEADER_ALIASES, MATERIAS_REQUIRED, _match_header_row

    src_ws = wb[template_name]
    new_ws = wb.copy_worksheet(src_ws)
    new_ws.title = new_name
    _copy_tables_from_source(wb, src_ws, new_ws)

    max_col = new_ws.max_column or 0
    max_row = new_ws.max_row or 0

    # 1) Materias: localiza cabecera y limpia filas de datos
    materias_header = None
    for r in range(1, min(max_row, 200) + 1):
        cols = _match_header_row(
            new_ws, r, MATERIAS_HEADER_ALIASES, MATERIAS_REQUIRED,
            extra_min_matches=2,
            extras_pool={"origen", "universidad_origen", "cuat", "firmado"},
        )
        if cols:
            materias_header = r
            break

    if materias_header:
        _clear_rows_preserve_style(new_ws, materias_header + 1, max_row, max_col)

    # 2) Catálogo: localiza cabecera y limpia filas de datos
    cat_info = find_catalog_in_ws(new_ws)
    if cat_info and cat_info["data_end"] > cat_info["header_row"]:
        _clear_rows_preserve_style(
            new_ws,
            cat_info["header_row"] + 1,
            cat_info["data_end"],
            max_col,
        )
        cat_info["data_end"] = cat_info["header_row"]

    # 3) Ajustar rangos de los openpyxl.Table a solo cabecera + 1 fila vacía
    for tbl in new_ws.tables.values():
        try:
            min_col, min_row, max_col_t, _ = range_boundaries(tbl.ref)
            tbl.ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col_t)}{min_row + 1}"
            )
        except Exception as e:
            logger.debug("No se pudo ajustar rango de Table '%s': %s", getattr(tbl, "name", "?"), e)


def find_materias_header_in_ws(ws) -> tuple[int, dict] | None:
    """
    Localiza la cabecera de la tabla de materias (Asignatura + Estudiante + ...)
    en la hoja. Devuelve (header_row_1based, cols_map) o None.
    """
    from ._excel_tables import (
        MATERIAS_HEADER_ALIASES, MATERIAS_REQUIRED, _match_header_row,
    )
    max_scan = min(ws.max_row or 0, 200)
    for r in range(1, max_scan + 1):
        cols_map = _match_header_row(
            ws, r, MATERIAS_HEADER_ALIASES, MATERIAS_REQUIRED,
            extra_min_matches=2,
            extras_pool={"origen", "universidad_origen", "cuat", "firmado"},
        )
        if cols_map:
            return r, cols_map
    return None


def insert_materias_rows(
    ws, header_row: int, cols_map: dict, rows_to_add: list[dict]
) -> int:
    """
    Inserta filas bajo la cabecera de materias, copiando estilo de la fila
    siguiente a la cabecera (que en hojas recién clonadas sigue conservando
    el formato de la primera fila de datos del curso original).
    Devuelve el índice 1-based de la última fila insertada.
    """
    c_asig = cols_map.get("asignatura")
    c_est  = cols_map.get("estudiante")
    c_ori  = cols_map.get("origen")
    c_uni  = cols_map.get("universidad_origen")
    c_cuat = cols_map.get("cuat")
    c_fir  = cols_map.get("firmado")
    c_la   = cols_map.get("link_la")

    fmt_row = header_row + 1
    insert_start = header_row + 1

    # Limitamos la copia de estilos al rango de columnas de la propia tabla de
    # materias; si no, pisaríamos el formato de la tabla de catálogo que suele
    # estar a la derecha en la misma hoja.
    mat_cols = [v for v in cols_map.values() if isinstance(v, int) and v > 0]
    col_min, col_max = (min(mat_cols), max(mat_cols)) if mat_cols else (1, ws.max_column or 1)

    for i, fila in enumerate(rows_to_add):
        r = insert_start + i
        for col_idx in range(col_min, col_max + 1):
            src = ws.cell(row=fmt_row, column=col_idx)
            dst = ws.cell(row=r, column=col_idx)
            if src.has_style:
                dst._style = copy(src._style)
        if c_asig: ws.cell(row=r, column=c_asig).value = fila["Asignatura"]
        if c_est:  ws.cell(row=r, column=c_est).value  = fila["Estudiante"]
        if c_ori:  ws.cell(row=r, column=c_ori).value  = fila["Origen"]
        if c_uni:  ws.cell(row=r, column=c_uni).value  = fila["Universidad Origen"]
        if c_cuat: ws.cell(row=r, column=c_cuat).value = fila["Cuat"]
        if c_fir:  ws.cell(row=r, column=c_fir).value  = fila["Firmado"]
        if c_la:   ws.cell(row=r, column=c_la).value   = fila["LA"]

    return insert_start + len(rows_to_add) - 1


def extend_tables_ref_to_row(ws, last_row: int) -> None:
    """Extiende el rango de cada openpyxl.Table de la hoja hasta `last_row`."""
    for tbl in ws.tables.values():
        try:
            mc, mr, Mc, _ = range_boundaries(tbl.ref)
            tbl.ref = f"{get_column_letter(mc)}{mr}:{get_column_letter(Mc)}{last_row}"
        except Exception as e:
            logger.debug("No se pudo extender Table '%s': %s", getattr(tbl, "name", "?"), e)
