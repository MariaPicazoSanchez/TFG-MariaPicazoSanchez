from __future__ import annotations
import os
import re
import logging
import unicodedata
import streamlit as st
from domain.models import ESTADOS_FIRMA
from utils import open_in_system
import pycountry
from babel import Locale
from persistence.data_access_mobility import get_universities_from_coords_sheet, get_universities_from_sicue_data
import pandas as pd
def get_university_country_map(path: str) -> dict:
    """
    Returns a dict {university: country} from the 'Coordenadas' sheet of the Excel file.
    """
    try:
        df_coords = pd.read_excel(path, sheet_name="Coordenadas", header=None, dtype=str)
        # col0: país, col1: universidad
        df_coords = df_coords.dropna(subset=[0, 1])
        return {str(row[1]).strip(): str(row[0]).strip() for _, row in df_coords.iterrows()}
    except Exception:
        return {}

def get_university_responsable_map(path: str) -> dict:
    """
    Returns {university: responsable} from the 'Coordenadas' sheet (col1/col3).

    Si la columna de responsable (col3/D) ya existe y tiene datos, se usa directamente.
    Si no existe o está vacía, escanea las hojas de alumnos del Excel para construir
    el mapa {universidad → responsable} y lo escribe en col3 de Coordenadas para
    que la próxima vez se lea directamente de ahí.
    """
    if not path:
        return {}
    try:
        df_coords = pd.read_excel(path, sheet_name="Coordenadas", header=None, dtype=str)

        # ── 1. Intentar leer de col3 (columna D) ──────────────────────────────
        if df_coords.shape[1] >= 4:
            result = {}
            for _, row in df_coords.iterrows():
                uni  = str(row.iloc[1] if pd.notna(row.iloc[1]) else "").strip()
                resp = str(row.iloc[3] if pd.notna(row.iloc[3]) else "").strip()
                if uni and resp and resp.lower() not in ("nan", "none", ""):
                    result[uni] = resp
            if result:
                return result

        # ── 2. Construir el mapa escaneando hojas de alumnos ─────────────────
        resp_map = _build_responsable_from_students(path)
        if resp_map:
            _write_responsable_to_coordenadas(path, resp_map)
        return resp_map

    except Exception:
        return {}


def _build_responsable_from_students(path: str) -> dict:
    """
    Recorre las hojas de datos del Excel buscando columnas de universidad
    y responsable, y construye el mapa {universidad: responsable}.
    La primera ocurrencia por universidad es la que se conserva.
    """
    import re

    def _is_academic_sheet(name: str) -> bool:
        return bool(re.search(r'\d{4}|\d{2}[-/]\d{2}', str(name)))

    def _pick_col(df: pd.DataFrame, *names: str):
        norm = lambda s: s.strip().lower()
        for n in names:
            for c in df.columns:
                if norm(str(c)) == norm(n):
                    return c
        return None

    resp_map: dict = {}
    try:
        wb_sheets = pd.ExcelFile(path, engine="openpyxl").sheet_names
    except Exception:
        return {}

    hojas = [s for s in wb_sheets if _is_academic_sheet(s)]
    if not hojas:
        hojas = [s for s in wb_sheets if s.lower() != "coordenadas"]

    for sheet in hojas:
        try:
            df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl", dtype=str)
            df.columns = [str(c).strip() for c in df.columns]

            c_uni  = _pick_col(df, "Destino", "destino", "Universidad", "universidad")
            c_resp = _pick_col(
                df,
                "Responsable programa", "responsable programa",
                "Responsable del programa", "responsable del programa",
                "Responsable", "responsable",
            )

            if not c_uni or not c_resp:
                continue

            for _, row in df.iterrows():
                uni  = str(row.get(c_uni,  "") or "").strip()
                resp = str(row.get(c_resp, "") or "").strip()
                if (uni and resp
                        and uni.lower()  not in ("nan", "none", "")
                        and resp.lower() not in ("nan", "none", "")):
                    resp_map.setdefault(uni, resp)  # primera ocurrencia gana

        except Exception:
            continue

    return resp_map


def _write_responsable_to_coordenadas(path: str, resp_map: dict) -> None:
    """
    Escribe el responsable de cada universidad en la columna D (col3)
    de la hoja 'Coordenadas', creando la celda si no existe.
    """
    import logging as _logging
    _log = _logging.getLogger("movilidad_ui")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        if "Coordenadas" not in wb.sheetnames:
            return
        ws = wb["Coordenadas"]
        for r_idx in range(1, ws.max_row + 1):
            uni_val = str(ws.cell(row=r_idx, column=2).value or "").strip()
            if uni_val and uni_val.lower() not in ("nan", "none", ""):
                resp_val = resp_map.get(uni_val, "")
                if resp_val:
                    ws.cell(row=r_idx, column=4).value = resp_val
        wb.save(path)
        _log.info("Responsables escritos en hoja Coordenadas: %d universidades", len(resp_map))
    except Exception as e:
        _log.warning("No se pudo escribir responsable en Coordenadas: %s", e)
from constants import PROGRAM_ERASMUS_IN, PROGRAM_ERASMUS_OUT
from constants import PROGRAM_ERASMUS_OUT, PROGRAM_ERASMUS_IN, PROGRAM_SICUE_OUT
from domain.validators import (DataValidator, safe_int_convert, is_duration_valid)
from persistence import get_asignaturas_catalog

from .sidebar import pick_local_file
from utils.app_config import save_course
USE_LOCAL_PICKER = True

logger = logging.getLogger("movilidad_ui")


def _clear_new_user_form_state():
    """
    Limpia TODO el formulario de nuevo usuario.
    """
    # 1) Borrar todos los campos del formulario (nu_*)
    for k in list(st.session_state.keys()):
        if k.startswith("nu_"):
            del st.session_state[k]

    # 2) Seguridad extra: forzar vacío en estos tres por si Streamlit
    # los vuelve a crear con valor anterior en este mismo run
    for k in ("nu_nombre", "nu_apellidos", "nu_email", "nu_destino_origen", "nu_pais_out", "nu_ciudad", "nu_tor", "nu_curso","nu_la_out_opt", "nu_acta", "nu_dur_out", "nu_resp_prog", "nu_plan_out", "nu_pais_in", "nu_la_in", "nu_horario", "nu_cuatri_in", "nu_la_sicue", "nu_plan", "nu_dur_sicue", "nu_coord_dest", "nu_materias_in", "nu_plan_sic_out"):
        st.session_state[k] = ""
    
    st.session_state["nu_firmado_la"] = False
    st.session_state["nu_investigacion_in"] = False
    st.session_state["_nu_inv_stable"] = False
    st.session_state["nu_estado"] = ESTADOS_FIRMA[0] if ESTADOS_FIRMA else ""

# ────────────────────────────────────────────────────────────────────────────────
# Lista de países (estándar ISO) con caché
# ────────────────────────────────────────────────────────────────────────────────

import streamlit as st
import pycountry
from babel import Locale

COUNTRY_ALIASES = {
    "Chequia": "República Checa",
    "Eslovaquia": "República Eslovaca",
    "Corea del Sur": "Corea, República de",
    "Corea del Norte": "Corea, República Popular Democrática de",
}

@st.cache_data
def get_country_options() -> list[str]:

    locale_es = Locale("es")
    nombres = []

    for country in pycountry.countries:

        nombre_es = locale_es.territories.get(country.alpha_2, country.name)

        # aplicar alias si existe
        nombre_es = COUNTRY_ALIASES.get(nombre_es, nombre_es)

        nombres.append(nombre_es)

    return [""] + sorted(set(nombres))


COUNTRY_OPTIONS = get_country_options()


def _load_asignaturas_catalog(config: dict, sheet_name: str | None = None) -> list:
    """Carga el catálogo de asignaturas de una hoja concreta, con caché en session_state."""
    ruta = (config.get("Erasmus IN") or "").strip()
    cache_key = f"_asignaturas_catalog_cache_{ruta}_{sheet_name or ''}"
    cached = st.session_state.get(cache_key)
    
    if cached and isinstance(cached, list) and len(cached) > 0 and "matriculados" not in cached[0]:
        del st.session_state[cache_key]
        
    if cache_key not in st.session_state:
        try:
            from persistence import get_asignaturas_catalog
            # ELIMINADO: sheet_name = st.session_state.get("global_sheet") <-- Esta línea causaba el bug
            st.session_state[cache_key] = get_asignaturas_catalog(config, sheet_name=sheet_name)
        except Exception as e:
            logger.warning("No se pudo cargar catálogo de asignaturas: %s", e)
            st.session_state[cache_key] = []
            
    return st.session_state.get(cache_key, [])




# ────────────────────────────────────────────────────────────────────────────────
# Geocoding (con caché)
# ────────────────────────────────────────────────────────────────────────────────
try:
    from geopy.geocoders import Nominatim
    from geopy.extra.rate_limiter import RateLimiter
    _GEOCODER = Nominatim(user_agent="tfg-movilidad-esii")
    _GEOCODE = RateLimiter(_GEOCODER.geocode, min_delay_seconds=1)
except Exception:
    _GEOCODER = None
    _GEOCODE = None


def _geocode_cached(q: str) -> tuple[float | None, float | None, str | None]:
    if not q or not q.strip():
        return None, None, "Consulta vacía"
    qn = q.strip().lower()
    cache = st.session_state.setdefault("_geo_cache", {})
    if qn in cache:
        d = cache[qn]
        return d["lat"], d["lon"], None
    if _GEOCODE is None:
        return None, None, "Geocoder no disponible"
    try:
        loc = _GEOCODE(q, addressdetails=False, timeout=10)
        if loc:
            lat, lon = float(loc.latitude), float(loc.longitude)
            cache[qn] = {"lat": lat, "lon": lon}
            return lat, lon, None
        return None, None, "No encontrado"
    except Exception as e:
        return None, None, f"Error geocodificando: {e}"



# ────────────────────────────────────────────────────────────────────────────────
# UI: formulario
# ────────────────────────────────────────────────────────────────────────────────
def _is_academic_year(name: str) -> bool:
    """Devuelve True si el nombre parece un curso académico (ej: 25-26, 2025/2026, 2016)."""
    return bool(re.search(r'\d{4}|\d{2}[-/]\d{2}', name))

def _sheet_options_for(cfg: dict, tipo: str) -> list[str]:
    from persistence import sheets_for
    sheets_map = (cfg or {}).get("sheets", {}) or {}
    known = sheets_map.get(tipo) or []
    if known:
        return sorted({s for s in known if s and s != "__CSV__" and _is_academic_year(s)})
    path = (cfg or {}).get(tipo, "")
    return [s for s in sheets_for(path) if s != "__CSV__" and _is_academic_year(s)] if path else []

def _invisible_suffix_from_id(button_id: str) -> str:
    """
    Genera un sufijo invisible único a partir de button_id,
    usando combinaciones de caracteres zero-width.
    """
    # Codificamos el button_id en bits y los traducimos a caracteres invisibles.
    bits = "".join(f"{ord(c):08b}" for c in button_id)  # p.ej. '01101010...'
    return "".join("\u200b" if b == "0" else "\u200c" for b in bits)
    # \u200b y \u200c son invisibles, pero la secuencia será distinta para cada id.

def _normalize_subject_name(s: str) -> str:
    """Normaliza nombre de asignatura para comparar: trim, unifica espacios,
    quita acentos y casefold para comparación insensible a mayúsc/minúsc."""
    s = (s or "").strip()
    s = re.sub(r'\s+', ' ', s)
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    return s.casefold()

_FILTER_ALL      = None  # cualquier archivo (LA)
_FILTER_PDF_WORD = "PDF y Word|*.pdf;*.docx|Todos|*.*"

def file_picker_button(label: str, text_input_key: str, button_id: str,
                       help: str = "Seleccionar archivo del equipo",
                       file_filter: str | None = _FILTER_PDF_WORD):
    invisible_suffix = _invisible_suffix_from_id(button_id)
    real_label = label + invisible_suffix

    clicked = st.button(real_label, help=help, key=button_id)

    if clicked:
        if USE_LOCAL_PICKER:
            current_val = st.session_state.get(text_input_key, "")
            path = pick_local_file(current_val, file_filter=file_filter)
            if path:
                # Guardar en clave buffer (no-widget) para que Streamlit no la limpie
                # al hacer rerun antes de que el text_input se renderice
                st.session_state["_buf_" + text_input_key] = path
        else:
            st.sidebar.warning("Ejecuta la app en local para seleccionar rutas del equipo.")
        st.rerun()
    return clicked



def _asig_label(a: dict) -> str:
    """Formatea una asignatura del catálogo con su info de matriculados/cupo."""
    nombre = a["asignatura"]
    matr = a.get("matriculados")
    cupo = a.get("cupo")
    if matr is not None and cupo is not None:
        return f"{nombre}  ({matr}/{cupo})"
    if matr is not None:
        return f"{nombre}  ({matr} matriculados)"
    return nombre


def _asig_nombre_puro(label: str) -> str:
    """Extrae solo el nombre de asignatura quitando el sufijo de matriculados/cupo."""
    return label.split("  (")[0].strip() if "  (" in label else label.strip()


def render_new_user_form(available_types: list[str], config: dict) -> dict | None:
    from domain import ESTADOS_FIRMA, ICON_BY_TIPO, CITIES_ES
    from persistence import append_user_to_excel, first_sheet_name

    # Transferir valores de buffer (file picker) a claves de widget ANTES de renderizar
    # Esto evita que Streamlit limpie el widget state al hacer rerun desde file_picker_button
    for _buf_key in [k for k in st.session_state if k.startswith("_buf_nu_")]:
        _widget_key = _buf_key[len("_buf_"):]
        st.session_state[_widget_key] = st.session_state.pop(_buf_key)

    # Si venimos de un guardado correcto:
    if st.session_state.pop("_user_saved", False):
        _clear_new_user_form_state()

        st.toast("Guardado correctamente", icon="✅")
    st.header("👤 Crear nuevo estudiante")

    if not available_types:
        st.warning("No hay ficheros Excel cargados. No puedes crear estudiantes todavía.")
        if st.button("🔁 Abrir ‘Cambiar rutas’"):
            st.session_state["show_routes"] = True
            st.session_state["view"] = "map"
            st.rerun()
        return None

    cfg = st.session_state.get("config", {}) or {}

    # Selectores en la misma fila
    col_tipo, col_sheet = st.columns([1, 1], gap="small")

    with col_tipo:
        prev_tipo = st.session_state.get("new_user_tipo", available_types[0])
        if prev_tipo not in available_types:
            prev_tipo = available_types[0]

        tipo = st.selectbox(
            "Tipo de alumno",
            options=available_types,
            index=available_types.index(prev_tipo),
            key="new_user_tipo",
        )
        tipo_norm = tipo
        open_label = f"{ICON_BY_TIPO.get(tipo_norm, '📄')} Abrir {tipo_norm}"

    with col_sheet:
        sheet_opts = _sheet_options_for(cfg, tipo_norm)
        SENT_NEW = "➕ Nueva hoja…"
        options = ([SENT_NEW] + sheet_opts) if sheet_opts else [SENT_NEW]

        # índice por defecto
        prev_sheet = st.session_state.get("new_user_sheet")
        global_sel = st.session_state.get("global_sheet", "Todas")
        if prev_sheet in options:
            idx = options.index(prev_sheet)
        elif global_sel in options:
            idx = options.index(global_sel)
        else:
            idx = 1 if len(options) > 1 else 0

        choice = st.selectbox("Curso", options=options, index=idx, key="new_user_sheet")
        new_sheet_name = None
        if choice == SENT_NEW:
            new_sheet_name = st.text_input("Nombre de la nueva hoja", key="nu_sheet_new_name", placeholder="2025-2026")

    selected_sheet = (new_sheet_name.strip() if new_sheet_name else (None if choice == SENT_NEW else choice))
    st.session_state["nu_sheet"] = selected_sheet

    # Guardar curso seleccionado en config.json
    if selected_sheet and selected_sheet != st.session_state.get("global_sheet"):
        save_course(selected_sheet)
        st.session_state["global_sheet"] = selected_sheet

    # Cargar catálogo de asignaturas para la hoja seleccionada
    asignaturas_catalog = _load_asignaturas_catalog(cfg, sheet_name=selected_sheet)

    # ────────────────────────────────────────────────────────────────
    # FORMULARIO PRINCIPAL (SIN st.form para permitir dinámica de asignaturas)
    # ────────────────────────────────────────────────────────────────
    
    # Inicializar variables comunes
    nombre = apellidos = destino_origen = email = ""
    extra: dict = {}
    st.markdown("""
    <style>
    /* selector genérico que intenta afectar solo a botones cuyo texto empieza por el icono carpeta */
    button[aria-label^="📁"] {
        height: 38px !important;
        width: 38px !important;
        padding: 0 !important;
    }
    </style>
    """, unsafe_allow_html=True)
    # _____________________________________
    # Erasmus OUT
    # _____________________________________
    if tipo_norm == PROGRAM_ERASMUS_OUT:
        xlsx_path = config.get(PROGRAM_ERASMUS_OUT, "")
        universidades_out = get_universities_from_coords_sheet(xlsx_path)
        uni_country_map_out = get_university_country_map(xlsx_path)
        resp_map_out = get_university_responsable_map(xlsx_path)
        with st.container(border=True):
            # Campos comunes
            col1, col2 = st.columns(2)
            with col1:
                nombre = st.text_input("Nombre", key="nu_nombre")
                email = st.text_input("Email", key="nu_email")
                # Autocompletar país si la universidad está en el mapeo
                pais_sugerido = uni_country_map_out.get(destino_origen.strip(), "") if destino_origen else ""
                # Si hay país sugerido y es distinto al actual, actualiza el campo país automáticamente
                if pais_sugerido and pais_sugerido in COUNTRY_OPTIONS:
                    if st.session_state.get("nu_pais_out", "") != pais_sugerido:
                        st.session_state["nu_pais_out"] = pais_sugerido
            with col2:
                apellidos = st.text_input("Apellidos", key="nu_apellidos")

                universidades_out_combo = [""] + universidades_out

                destino_origen = st.selectbox(
                    "Destino (universidad)",
                    options=universidades_out_combo,
                    key="nu_destino_origen",
                    help="Selecciona una universidad o escribe una nueva",
                    accept_new_options=True,
                )

                # Autocompletar país automáticamente
                if destino_origen:
                    pais_sugerido = uni_country_map_out.get(destino_origen.strip(), "")

                    if pais_sugerido and pais_sugerido in COUNTRY_OPTIONS:
                        if st.session_state.get("nu_pais_out") != pais_sugerido:
                            st.session_state["nu_pais_out"] = pais_sugerido
            
            # Campos específicos
            col1, col2 = st.columns(2)
            with col1:
                # El valor lo gestiona session_state, no usar index para evitar warning
                extra["pais_out"] = st.selectbox("País", options=COUNTRY_OPTIONS, key="nu_pais_out")
                dur_out_val = st.text_input("Duración (meses)", key="nu_dur_out")
                if dur_out_val and not dur_out_val.strip().isdigit():
                    st.toast("⚠️ La duración debe ser un número", icon="⚠️")
                    extra["dur_out"] = ""
                else:
                    extra["dur_out"] = dur_out_val
                # Validar que solo contiene números
                act_col1, act_col2 = st.columns([8, 1.5])
                with act_col2:
                    st.markdown("<br>", unsafe_allow_html=True)
                    file_picker_button("📁", "nu_acta", "nu_acta_browse", "Abrir explorador de archivos.")
                with act_col1:
                    extra["acta_equivalencias"] = st.text_input("Acta de equivalencias (ruta o enlace)", key="nu_acta")

                tor_col1, tor_col2 = st.columns([8, 1.5])
                with tor_col1:
                    extra["tor"] = st.text_input("ToR (ruta o enlace)", key="nu_tor")
                with tor_col2:
                    st.markdown("<br>", unsafe_allow_html=True)
                    file_picker_button("📁", "nu_tor", "nu_tor_browse", "Abrir explorador de archivos.")

                plan_col1, plan_col2 = st.columns([8, 1.5])
                with plan_col1:
                    extra["plan_out"] = st.text_input(
                        "Propuesta alumno LA (ruta o enlace)",
                        key="nu_plan_out",
                    )
                with plan_col2:
                    st.markdown("<br>", unsafe_allow_html=True)
                    file_picker_button("📁", "nu_plan_out", "nu_plan_out_browse", "Abrir explorador de archivos.")
                
                
            with col2:
                extra["curso"] = st.selectbox("Curso", options=["","1", "2", "3", "4"], key="nu_curso")
                extra["ciudad"] = st.text_input("Ciudad", key="nu_ciudad")
                resp_auto = resp_map_out.get((destino_origen or "").strip(), "")
                extra["resp_prog"] = resp_auto
                if resp_auto:
                    st.caption(f"Responsable: **{resp_auto}**")
                la_col1, la_col2 = st.columns([8, 1.5])
                with la_col1:
                    extra["la_out"] = st.text_input("LA (enlace o ruta)", key="nu_la_out_opt")
                with la_col2:
                    st.markdown("<br>", unsafe_allow_html=True)
                    file_picker_button("📁", "nu_la_out_opt", "nu_la_out_opt_browse", "Abrir explorador de archivos.", file_filter=_FILTER_ALL)
                
    
    # _____________________________________
    # Erasmus IN
    # _____________________________________

    elif tipo_norm == PROGRAM_ERASMUS_IN:
        xlsx_path = config.get(PROGRAM_ERASMUS_IN, "")
        universidades_in = get_universities_from_coords_sheet(xlsx_path)
        uni_country_map_in = get_university_country_map(xlsx_path)
        # Streamlit puede limpiar el widget state de nu_investigacion_in cuando
        # file_picker_button llama a st.rerun() antes de que el checkbox se renderice.
        # Por eso mantenemos una clave estable (_nu_inv_stable) que no es un widget
        # y que Streamlit no limpia automáticamente. Usamos el widget state si está
        # disponible (valor más reciente), y si no, recurrimos a la clave estable.
        es_investigacion = st.session_state.get(
            "nu_investigacion_in",
            st.session_state.get("_nu_inv_stable", False)
        )
        with st.container(border=True):
            col1, col2 = st.columns(2)

            with col2:
                apellidos = st.text_input("Apellidos", key="nu_apellidos")

                universidades_in_combo = [""] + universidades_in

                destino_origen = st.selectbox(
                    "Origen (universidad)",
                    options=universidades_in_combo,
                    key="nu_destino_origen",
                    help="Selecciona una universidad o escribe una nueva",
                    accept_new_options=True,
                )
                la_col1, la_col2 = st.columns([8, 1.5])
                with la_col2:
                    st.markdown("<br>", unsafe_allow_html=True)
                    file_picker_button("📁", "nu_la_in", "nu_la_in_browse", "Abrir explorador de archivos.", file_filter=_FILTER_ALL)
                with la_col1:
                    extra["la_in"] = st.text_input("LA (enlace o ruta)", key="nu_la_in")

                st.markdown("<div style='margin-top:30px'></div>", unsafe_allow_html=True)
                extra["investigacion_in"] = st.checkbox(
                    "Investigación",
                    key="nu_investigacion_in"
                )
                # Guardar en clave estable (no-widget) para sobrevivir al widget cleanup de Streamlit
                st.session_state["_nu_inv_stable"] = extra["investigacion_in"]

            # AUTOCOMPLETAR PAÍS (antes del selectbox)
            pais_sugerido = ""
            if destino_origen:
                pais_sugerido = uni_country_map_in.get(destino_origen.strip(), "")

            if pais_sugerido and pais_sugerido in COUNTRY_OPTIONS:
                if st.session_state.get("nu_pais_in") != pais_sugerido:
                    st.session_state["nu_pais_in"] = pais_sugerido
            with col1:
                nombre = st.text_input("Nombre", key="nu_nombre")

                extra["cuatrimestre_in"] = st.selectbox(
                    "Cuatrimestre",
                    options=["", "1", "2"],
                    key="nu_cuatri_in"
                )

                extra["pais_in"] = st.selectbox(
                    "País",
                    options=COUNTRY_OPTIONS,
                    key="nu_pais_in"
                )
                
                st.markdown("<div style='margin-top:30px'></div>", unsafe_allow_html=True)

                extra["firmado_la"] = "x" if st.checkbox(
                        "LA firmado",
                        key="nu_firmado_la"
                    ) else ""
            # ─────────────────────────────
            # ASIGNATURAS (YA DENTRO)
            # ─────────────────────────────

            if not es_investigacion:
                st.divider()
                st.markdown("#### 📚 Asignaturas")

                materias_key = "nu_materias_in"
                # Asegurar que materias siempre sea una lista
                if materias_key not in st.session_state or not isinstance(st.session_state[materias_key], list):
                    st.session_state[materias_key] = []
                materias = st.session_state[materias_key]

                cuatrimestre_seleccionado = st.session_state.get("nu_cuatri_in", "")

                if cuatrimestre_seleccionado:
                    asignaturas_sugerencias = [
                        a["asignatura"]
                        for a in asignaturas_catalog
                        if a.get("cuat") == cuatrimestre_seleccionado
                        and a["asignatura"].strip().lower() != "estancia investigación"
                    ]
                else:
                    asignaturas_sugerencias = [
                        a["asignatura"]
                        for a in asignaturas_catalog
                        if a["asignatura"].strip().lower() != "estancia investigación"
                    ]

                header_cols = st.columns([8, 1], vertical_alignment="bottom")

                # Mapa nombre -> info del catálogo para mostrar matriculados/cupo
                catalog_map = {a["asignatura"]: a for a in asignaturas_catalog}

                header_cols_h = st.columns([8, 2, 1], vertical_alignment="bottom")
                with header_cols_h[0]:
                    st.caption("Nombre de la asignatura")
                with header_cols_h[1]:
                    st.caption("Matr. / Cupo")
                with header_cols_h[2]:
                    if st.button("➕ Añadir", key=f"{materias_key}_add"):
                        materias.append({"nombre": ""})

                delete_idx = None

                for i, mat in enumerate(materias):
                    # Leer selección actual del session_state ANTES de renderizar columnas
                    raw_sel = st.session_state.get(f"{materias_key}_select_{i}")
                    nom_actual = _asig_nombre_puro(raw_sel) if raw_sel else _asig_nombre_puro(mat.get("nombre", ""))
                    info = catalog_map.get(nom_actual)

                    row_cols = st.columns([8, 2, 1], vertical_alignment="center")
                    with row_cols[0]:
                        valor_actual = mat.get("nombre", "")
                        seleccion = st.selectbox(
                            f"Asignatura {i+1}",
                            options=asignaturas_sugerencias,
                            index=asignaturas_sugerencias.index(valor_actual)
                                if valor_actual in asignaturas_sugerencias else None,
                            key=f"{materias_key}_select_{i}",
                            label_visibility="collapsed",
                            placeholder="Seleccionar o escribir...",
                            accept_new_options=True
                        )
                        mat["nombre"] = _asig_nombre_puro(seleccion) if seleccion else ""
                    with row_cols[1]:
                        if info:
                            matr = info.get("matriculados")
                            cupo = info.get("cupo")
                            # +1 porque este estudiante aún no está guardado
                            matr_display = (matr + 1) if matr is not None else None
                            if matr_display is not None and cupo is not None:
                                color = "#e05252" if matr_display > cupo else "#4caf50"
                                st.markdown(
                                    f"<p style='margin:-0.5rem 0 0 -0.1rem;font-size:1.1rem;font-weight:700;"
                                    f"color:{color};text-align:left;line-height:2.4rem'>"
                                    f"{matr_display}&nbsp;/&nbsp;{cupo}</p>",
                                    unsafe_allow_html=True
                                )
                            elif matr_display is not None:
                                st.markdown(
                                    f"<p style='margin:-0.5rem 0 0 -0.1rem;font-size:1.1rem;font-weight:700;"
                                    f"color:#888;text-align:left;line-height:2.4rem'>"
                                    f"{matr_display}</p>",
                                    unsafe_allow_html=True
                                )
                        else:
                            st.empty()
                    with row_cols[2]:
                        if st.button(
                            "❌",
                            key=f"{materias_key}_del_{i}",
                            help="Eliminar asignatura",
                            type="secondary",
                            use_container_width=True
                        ):
                            delete_idx = i

                if delete_idx is not None:
                    materias.pop(delete_idx)

                # ── Validación en tiempo real ──────────────────────────────────
                if not materias:
                    st.warning("Debes añadir al menos una asignatura antes de guardar el estudiante.")
                else:
                    nombres_rellenos = [(i, (m.get("nombre") or "").strip()) for i, m in enumerate(materias)]

                    # Filas vacías
                    vacias = [i + 1 for i, n in nombres_rellenos if not n]
                    if vacias:
                        fila_txt = ", ".join(f"#{f}" for f in vacias)
                        st.warning(
                            f"⚠️ {'La asignatura' if len(vacias) == 1 else 'Las asignaturas'} "
                            f"{fila_txt} {'está vacía' if len(vacias) == 1 else 'están vacías'}. "
                            f"Rellénala{'s' if len(vacias) > 1 else ''} o elimínala{'s' if len(vacias) > 1 else ''}."
                        )

                    # Duplicados (solo entre las que tienen nombre)
                    seen_norm: dict[str, int] = {}
                    duplicadas: list[str] = []
                    for i, n in nombres_rellenos:
                        if not n:
                            continue
                        nk = _normalize_subject_name(n)
                        if nk in seen_norm:
                            duplicadas.append(n)
                        else:
                            seen_norm[nk] = i + 1

                    if duplicadas:
                        dup_txt = ", ".join(sorted(set(duplicadas)))
                        st.error(
                            f"❌ Asignatura{'s' if len(set(duplicadas)) > 1 else ''} "
                            f"repetida{'s' if len(set(duplicadas)) > 1 else ''}: {dup_txt}"
                        )
    # _____________________________________
    # SICUE OUT
    # _____________________________________

    elif tipo_norm == PROGRAM_SICUE_OUT:
        xlsx_path_sicue = config.get(PROGRAM_SICUE_OUT, "")
        universidades_sicue, ciudad_map_sicue, coords_map_sicue = (
            get_universities_from_sicue_data(xlsx_path_sicue)
            if xlsx_path_sicue else ([], {}, {})
        )

        with st.container(border=True):
            # Campos comunes
            col1, col2 = st.columns(2)
            with col1:
                nombre = st.text_input("Nombre", key="nu_nombre")
                email = st.text_input("Email", key="nu_email")
            with col2:
                apellidos = st.text_input("Apellidos", key="nu_apellidos")

                universidades_sicue_combo = [""] + universidades_sicue
                destino_origen = st.selectbox(
                    "Destino (universidad)",
                    options=universidades_sicue_combo,
                    key="nu_destino_origen",
                    help="Selecciona una universidad ya conocida o escribe una nueva",
                    accept_new_options=True,
                )

                # Autocompletar ciudad si la universidad ya está en los datos
                if destino_origen:
                    ciudad_sugerida = ciudad_map_sicue.get(destino_origen.strip(), "")
                    if ciudad_sugerida and st.session_state.get("nu_ciudad", "") != ciudad_sugerida:
                        st.session_state["nu_ciudad"] = ciudad_sugerida

            # Guardar coords conocidas en session_state para usarlas al guardar
            coords_conocidas = coords_map_sicue.get((destino_origen or "").strip())
            st.session_state["_sicue_coords_known"] = coords_conocidas  # (lat, lon) o None

            # Campos específicos
            col1, col2 = st.columns(2)
            with col1:
                extra["ciudad_sicue"] = st.selectbox(
                    "Ciudad",
                    options=CITIES_ES,
                    index=0,
                    help="Si no se encuentran coordenadas por la universidad, se intentará con esta ciudad.",
                    key="nu_ciudad",
                )
                dur_sicue_val = st.text_input("Duración (meses)", key="nu_dur_sicue")
                if dur_sicue_val and not dur_sicue_val.strip().isdigit():
                    st.toast("La duración debe ser un número", icon="⚠️")
                    extra["dur_sicue"] = ""
                else:
                    extra["dur_sicue"] = dur_sicue_val
                la_col1, la_col2 = st.columns([8, 1.5])
                with la_col2:
                    st.markdown("<br>", unsafe_allow_html=True)
                    file_picker_button("📁", "nu_la_sicue", "nu_la_sicue_browse", "Abrir explorador de archivos.", file_filter=_FILTER_ALL)
                with la_col1:
                    extra["la_in"] = st.text_input("LA (enlace o ruta)", key="nu_la_sicue")

            with col2:
                extra["estado_firmas"]  = st.selectbox("Estado de firmas", ESTADOS_FIRMA, key="nu_estado")
                extra["coord_dest"] = st.text_input("Coordinador en destino", key="nu_coord_dest")

                plan_col1, plan_col2 = st.columns([8, 1.5])
                with plan_col2:
                    st.markdown("<br>", unsafe_allow_html=True)
                    file_picker_button("📁", "nu_plan_sic_out", "nu_plan_sic_out_browse", "Abrir explorador de archivos.")

                with plan_col1:
                    extra["plan_sic_out"] = st.text_input(
                        "Plan de estudios (ruta o enlace)",
                        key="nu_plan_sic_out"
                    )
                
    

    # Botones de acción alineados en la misma fila y ocupando todo el ancho
    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:

        open_clicked = st.button(open_label, key="open_xlsx_button", use_container_width=True)
    with col_btn2:
        submit_clicked = st.button("Guardar estudiante", key="submit_new_user", use_container_width=True, type="primary")

    if open_clicked:
        xlsx_for_tipo = config.get(tipo_norm)
        if xlsx_for_tipo:
            ok2, err2 = open_in_system(os.path.abspath(xlsx_for_tipo))
            if not ok2:
                st.warning(f"No se pudo abrir el archivo: {err2}")
        else:
            st.warning(f"No hay Excel configurado para '{tipo_norm}'.")
        st.stop()  # evitar validar/guardar cuando solo se quiso abrir

    if not submit_clicked:
        return None

    # ────────────────────────────────────────────────────────────────
    # VALIDACIONES
    # ────────────────────────────────────────────────────────────────
    validator = DataValidator()

    
    # Campos obligatorios con normalización
    nombre_val = (nombre or "").strip()
    destino_val = (destino_origen or "").strip()
    
    if not nombre_val:
        validator._add_error("nombre", "El nombre es obligatorio")
    else:
        validator.cleaned_data["nombre"] = nombre_val
    
    if not destino_val:
        validator._add_error("destino_origen", "El destino/universidad es obligatorio")
    else:
        validator.cleaned_data["destino_origen"] = destino_val
    
    # Campos opcionales (añadir a cleaned_data si existen)
    apellidos_val = (apellidos or "").strip()
    if apellidos_val:
        validator.cleaned_data["apellidos"] = apellidos_val
    
    email_val = (email or "").strip()
    if email_val:
        validator.cleaned_data["email"] = email_val
    
    # Validaciones específicas por tipo con normalización
    if tipo_norm == PROGRAM_ERASMUS_OUT:
        pais_val = (extra.get("pais_out") or "").strip()
        if not pais_val:
            validator._add_error("pais_out", "El país es obligatorio")
        else:
            validator.cleaned_data["pais_out"] = pais_val
            
        if extra.get("dur_out"):
            validator.validate_field("dur_out", extra["dur_out"], is_duration_valid(),
                                    normalizer=lambda x: str(safe_int_convert(x, default=0)))
            
    elif tipo_norm == PROGRAM_ERASMUS_IN:
        pais_val = (extra.get("pais_in") or "").strip()
        if not pais_val:
            validator._add_error("pais_in", "El país es obligatorio")
        else:
            validator.cleaned_data["pais_in"] = pais_val
            
    elif tipo_norm == PROGRAM_SICUE_OUT:
        ciudad_val = (extra.get("ciudad_sicue") or "").strip()
        if not ciudad_val:
            validator._add_error("ciudad_sicue", "La ciudad es obligatoria")
        else:
            validator.cleaned_data["ciudad_sicue"] = ciudad_val
            
        if extra.get("dur_sicue"):
            validator.validate_field("dur_sicue", extra["dur_sicue"], is_duration_valid(),
                                    normalizer=lambda x: str(safe_int_convert(x, default=0)))
    
    if not validator.is_valid():
        st.toast(f"{validator.get_error_messages()}", icon="❌")
        return None
    if tipo_norm == PROGRAM_ERASMUS_IN:
        materias_raw = st.session_state.get("nu_materias_in", []) or []
        # obtener nombres, respetando orden y quitando filas vacías
        nombres = [_asig_nombre_puro((m.get("nombre") or "").strip()) for m in materias_raw]
        nombres_no_vacios = [n for n in nombres if n]

        seen = {}
        duplicates = []  # guardará las apariciones duplicadas en forma legible
        for idx, n in enumerate(nombres_no_vacios):
            nk = _normalize_subject_name(n)
            if nk in seen:
                # añadimos la forma original (no normalizada) para mostrar al usuario
                duplicates.append(n)
            else:
                seen[nk] = idx

        if duplicates:
            # Mensaje claro y conciso
            validator._add_error(
                "materias_in",
                "Hay asignaturas repetidas: " + ", ".join(sorted(set(duplicates)))
            )
        else:
            # Guardar lista limpia en cleaned_data para usar después
            validator.cleaned_data["materias_in"] = nombres_no_vacios
    
    # Obtener datos normalizados
    clean_data = validator.get_clean_data()


    # ────────────────────────────────────────────────────────────────
    # GEOCODING (solo para SICUE OUT)
    # ────────────────────────────────────────────────────────────────
    lat, lon = None, None
    if tipo_norm == PROGRAM_SICUE_OUT:
        # Si ya tenemos coordenadas de registros anteriores, usarlas directamente
        coords_conocidas = st.session_state.pop("_sicue_coords_known", None)
        if coords_conocidas:
            lat, lon = coords_conocidas
        else:
            lat, lon, gerr = _geocode_cached(destino_origen.strip())
            if lat is None or lon is None:
                ciudad_opt = (extra.get("ciudad_sicue") or "").strip()
                if ciudad_opt:
                    lat, lon, gerr2 = _geocode_cached(ciudad_opt)
                    if gerr and not gerr2:
                        gerr = None
            if lat is None and lon is None:
                st.warning(f"No se pudo geocodificar '{destino_origen}'")
    # ────────────────────────────────────────────────────────────────
    # PAYLOAD + MATERIAS
    # ────────────────────────────────────────────────────────────────
    # Usar datos normalizados del validador - mucho más limpio
    # Mapear claves para SICUE OUT a las que espera el Excel
    if tipo_norm == PROGRAM_SICUE_OUT:
        duracion = extra.get("dur_sicue", "")
        sicue_payload = {
            "tipo": tipo_norm,
            "nombre": clean_data.get("nombre", ""),
            "apellidos": clean_data.get("apellidos", ""),
            "email": clean_data.get("email", ""),
            "destino_origen": clean_data.get("destino_origen", ""),
            "coordenadas": (lat, lon),
            "ciudad_sicue": extra.get("ciudad_sicue", ""),
            "duracion_meses": duracion,
            "dur_sicue": duracion,
            "coord_dest": extra.get("coord_dest", ""),
            "estado_firmas": extra.get("estado_firmas", ""),
            "la_in": extra.get("la_in", ""),
            "firmado_la": extra.get("firmado_la", ""),
            "plan_sic_out": extra.get("plan_sic_out", ""),
            # Alias para compatibilidad con columnas del Excel
            "la": extra.get("la_in", ""),
            "gestion_la": extra.get("firmado_la", ""),
            "plan_estudios": extra.get("plan_sic_out", ""),
        }
        payload = sicue_payload
    else:
        payload = {
            "tipo": tipo_norm,
            "nombre": clean_data.get("nombre", ""),
            "apellidos": clean_data.get("apellidos", ""),
            "email": clean_data.get("email", ""),
            "destino_origen": clean_data.get("destino_origen", ""),
            "coordenadas": (lat, lon),
            **{k: (v.strip() if isinstance(v, str) else v) for k, v in extra.items()},
        }
    if tipo_norm == "Erasmus IN":
        materias_payload = []

        # Obtener valores globales del estudiante para todas las asignaturas
        cuat_global = extra.get("cuatrimestre_in", "")
        firmado_global = extra.get("firmado_la", "")
        la_global = extra.get("la_in", "")
        es_investigacion = extra.get("investigacion_in", False)

        if es_investigacion:
            materias_payload = [{
                "asignatura": "Estancia Investigación",
                "cuat": cuat_global,
                "firmado": firmado_global,
                "link_la": la_global
            }]
        else:
            for m in st.session_state.get("nu_materias_in", []):
                nom = _asig_nombre_puro((m.get("nombre") or "").strip())
                if not nom:
                    # ignoramos filas vacías
                    continue
                materias_payload.append({
                    "asignatura": nom,
                    "cuat": cuat_global,
                    "firmado": firmado_global,
                    "link_la": la_global
                })

            # Validar que hay al menos una asignatura
            if not materias_payload:
                st.toast("Debes añadir al menos una asignatura para Erasmus IN", icon="❌")
                return None

        if materias_payload:
            payload["materias_in"] = materias_payload
    else:
        materias_payload = []


    # ────────────────────────────────────────────────────────────────
    # GUARDAR EN EXCEL
    # ────────────────────────────────────────────────────────────────
    xlsx_path = config.get(tipo_norm)
    if not xlsx_path:
        st.toast(f"No hay Excel configurado para '{tipo_norm}'. Ábrelo en 'Cambiar rutas'.", icon="❌")
        return None

    ok, err = append_user_to_excel(xlsx_path, tipo_norm, payload, sheet_name=selected_sheet)
    if not ok:
        st.toast(f"Error guardando en Excel: {err}", icon="❌")
        return None

    # Marcamos éxito y forzamos incremento de `data_version` para invalidar caches
    st.session_state["_user_saved"] = True
    st.session_state["data_version"] = st.session_state.get("data_version", 0) + 1
    st.rerun()