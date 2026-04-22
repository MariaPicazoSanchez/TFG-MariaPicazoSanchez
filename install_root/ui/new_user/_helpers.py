"""
Helpers para el formulario de nuevo estudiante:
  - Mapas universidad→país, universidad→responsable
  - Lista de países (ISO con babel)
  - Catálogo de asignaturas
  - Geocodificación con caché
  - Utilidades de texto y selección de fichero
"""


import logging
import re
import unicodedata

import pandas as pd
import streamlit as st

logger = logging.getLogger("movilidad_ui")

# ──────────────────────────────────────────────────────────────────────────────
# Utilidades de selección de fichero
# ──────────────────────────────────────────────────────────────────────────────

USE_LOCAL_PICKER = True
_FILTER_ALL = None
_FILTER_PDF_WORD = "PDF y Word|*.pdf;*.docx|Todos|*.*"


def _invisible_suffix_from_id(button_id: str) -> str:
    """Genera un sufijo invisible único para evitar claves duplicadas en st.button."""
    bits = "".join(f"{ord(c):08b}" for c in button_id)
    return "".join("\u200b" if b == "0" else "\u200c" for b in bits)


def file_picker_button(
    label: str,
    text_input_key: str,
    button_id: str,
    help: str = "Seleccionar archivo del equipo",
    file_filter: str | None = _FILTER_PDF_WORD,
) -> bool:
    from ui.sidebar import pick_local_file

    invisible_suffix = _invisible_suffix_from_id(button_id)
    clicked = st.button(label + invisible_suffix, help=help, key=button_id)
    if clicked:
        if USE_LOCAL_PICKER:
            current_val = st.session_state.get(text_input_key, "")
            path = pick_local_file(current_val, file_filter=file_filter)
            if path:
                st.session_state["_buf_" + text_input_key] = path
        else:
            st.sidebar.warning("Ejecuta la app en local para seleccionar rutas del equipo.")
        st.rerun()
    return clicked


# ──────────────────────────────────────────────────────────────────────────────
# Detección de columnas en la hoja de Coordenadas
# ──────────────────────────────────────────────────────────────────────────────

def _detect_coords_columns(
    df: "pd.DataFrame",
    default_col_pais: int = 0,
    default_col_uni: int = 1,
) -> tuple[int, int, bool]:
    """
    Detecta las columnas País y Universidad en la hoja de Coordenadas.

    Si la primera fila contiene etiquetas de cabecera reconocibles ('universidad',
    'país', etc.), las usa para determinar los índices de columna y devuelve
    skip_first_row=True. En caso contrario, devuelve los valores por defecto.

    Returns:
        (col_pais, col_uni, skip_first_row)
    """
    _WORDS_UNI = {
        "universidad", "universidade", "university",
        "universidad destino", "universidad origen",
    }
    _WORDS_PAI = {
        "país", "pais", "country",
        "país/región", "pais/region", "país / región",
    }

    if len(df) == 0:
        return default_col_pais, default_col_uni, False

    first_row = [str(v).strip().lower() for v in df.iloc[0]]

    col_uni_det: int | None = None
    col_pai_det: int | None = None

    for i, v in enumerate(first_row):
        if v in _WORDS_UNI and col_uni_det is None:
            col_uni_det = i
        elif v in _WORDS_PAI and col_pai_det is None:
            col_pai_det = i

    is_header = (col_uni_det is not None) or (col_pai_det is not None)

    if is_header:
        return (
            col_pai_det if col_pai_det is not None else default_col_pais,
            col_uni_det if col_uni_det is not None else default_col_uni,
            True,
        )

    return default_col_pais, default_col_uni, False


# ──────────────────────────────────────────────────────────────────────────────
# Mapas universidad → país / responsable
# ──────────────────────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def get_university_country_map(
    path: str,
    default_col_uni: int = 1,
    default_col_pais: int = 0,
) -> dict:
    """
    Devuelve {universidad: país} leyendo la hoja 'Coordenadas'.

    Formato por defecto (Erasmus IN): col0=País, col1=Universidad.
    Formato Erasmus OUT:              col0=Universidad, col1=País.

    Si la primera fila contiene etiquetas de cabecera, se detectan las columnas
    automáticamente y esa fila se descarta de los datos.
    """
    _BAD = {"nan", "none", ""}
    try:
        df = pd.read_excel(path, sheet_name="Coordenadas", header=None, dtype=str)
        col_pais, col_uni, skip = _detect_coords_columns(df, default_col_pais, default_col_uni)
        if skip:
            df = df.iloc[1:].reset_index(drop=True)
        result = {}
        for _, row in df.iterrows():
            uni  = str(row.iloc[col_uni]).strip()
            pais = str(row.iloc[col_pais]).strip()
            if uni.lower() not in _BAD and pais.lower() not in _BAD:
                result[uni] = pais
        return result
    except Exception:
        return {}


@st.cache_data(show_spinner=False)
def get_university_responsable_map(
    path: str,
    default_col_uni: int = 1,
    default_col_pais: int = 0,
) -> dict:
    """
    Devuelve {universidad: responsable} desde la hoja 'Coordenadas' (col índice 3).
    Si la columna no existe, escanea las hojas de alumnos y escribe el resultado en col3.

    Formato por defecto (Erasmus IN): col0=País, col1=Universidad.
    Formato Erasmus OUT:              col0=Universidad, col1=País.

    Si la primera fila contiene etiquetas de cabecera, se detectan las columnas
    automáticamente y esa fila se descarta de los datos.
    """
    if not path:
        return {}
    try:
        df_coords = pd.read_excel(path, sheet_name="Coordenadas", header=None, dtype=str)
        col_pais, col_uni, skip = _detect_coords_columns(df_coords, default_col_pais, default_col_uni)
        if skip:
            df_coords = df_coords.iloc[1:].reset_index(drop=True)
        if df_coords.shape[1] >= 4:
            unis  = df_coords.iloc[:, col_uni].fillna("").astype(str).str.strip()
            resps = df_coords.iloc[:, 3].fillna("").astype(str).str.strip()
            _bad  = {"nan", "none", ""}
            mask  = unis.ne("") & resps.ne("") & ~resps.str.lower().isin(_bad)
            result = dict(zip(unis[mask], resps[mask]))
            if result:
                return result
        resp_map = _build_responsable_from_students(path)
        if resp_map:
            # col_uni es índice 0-based; openpyxl usa 1-based → col_uni + 1
            _write_responsable_to_coordenadas(path, resp_map, col_uni_openpyxl=col_uni + 1)
        return resp_map
    except Exception:
        return {}


def _build_responsable_from_students(path: str) -> dict:
    """Construye {universidad: responsable} escaneando las hojas de datos del Excel."""
    def _is_academic_sheet(name: str) -> bool:
        return bool(re.search(r'\d{4}|\d{2}[-/]\d{2}', str(name)))

    def _pick_col(df: pd.DataFrame, *names: str):
        norm = lambda s: s.strip().lower()
        for n in names:
            for c in df.columns:
                if norm(str(c)) == norm(n):
                    return c
        return None

    resp_map: dict = {}
    try:
        wb_sheets = pd.ExcelFile(path, engine="openpyxl").sheet_names
    except Exception:
        return {}

    hojas = [s for s in wb_sheets if _is_academic_sheet(s)]
    if not hojas:
        hojas = [s for s in wb_sheets if s.lower() != "coordenadas"]

    for sheet in hojas:
        try:
            df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl", dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
            c_uni  = _pick_col(df, "Destino", "destino", "Universidad", "universidad")
            c_resp = _pick_col(
                df,
                "Responsable programa", "responsable programa",
                "Responsable del programa", "responsable del programa",
                "Responsable", "responsable",
            )
            if not c_uni or not c_resp:
                continue
            unis  = df[c_uni].fillna("").astype(str).str.strip()
            resps = df[c_resp].fillna("").astype(str).str.strip()
            _bad  = {"nan", "none", ""}
            mask  = (unis.ne("") & resps.ne("")
                     & ~unis.str.lower().isin(_bad)
                     & ~resps.str.lower().isin(_bad))
            for uni, resp in zip(unis[mask].tolist(), resps[mask].tolist()):
                resp_map.setdefault(uni, resp)
        except Exception:
            continue
    return resp_map


def _write_responsable_to_coordenadas(
    path: str,
    resp_map: dict,
    col_uni_openpyxl: int = 2,
) -> None:
    """
    Escribe el responsable en col D (columna 4, 1-based) de la hoja 'Coordenadas'.

    col_uni_openpyxl: columna 1-based (openpyxl) donde está la universidad.
        Erasmus IN  → 2  (col B, por defecto)
        Erasmus OUT → 1  (col A)
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        if "Coordenadas" not in wb.sheetnames:
            return
        ws = wb["Coordenadas"]
        for r_idx in range(1, ws.max_row + 1):
            uni_val = str(ws.cell(row=r_idx, column=col_uni_openpyxl).value or "").strip()
            if uni_val and uni_val.lower() not in ("nan", "none", ""):
                resp_val = resp_map.get(uni_val, "")
                if resp_val:
                    ws.cell(row=r_idx, column=4).value = resp_val
        wb.save(path)
        logger.info("Responsables escritos en hoja Coordenadas: %d universidades", len(resp_map))
    except Exception as e:
        logger.warning("No se pudo escribir responsable en Coordenadas: %s", e)


# ──────────────────────────────────────────────────────────────────────────────
# Mapa universidad → plan de estudios (col índice 4 → col E, 1-based)
# ──────────────────────────────────────────────────────────────────────────────

PLAN_ESTUDIOS_COL_IDX = 4      # 0-based (para pandas)
PLAN_ESTUDIOS_COL_XLSX = 5     # 1-based (para openpyxl)


@st.cache_data(show_spinner=False)
def get_university_plan_estudios_map(
    path: str,
    default_col_uni: int = 1,
    default_col_pais: int = 0,
) -> dict:
    """
    Devuelve {universidad: plan_estudios} desde la hoja 'Coordenadas' (col índice 4).
    Si la columna no existe o está vacía, escanea las hojas de alumnos y escribe el
    resultado en col E para que quede persistido.

    Mismo comportamiento que get_university_responsable_map pero para la columna
    'Plan de estudios' (col E, índice 4 basado en 0 / 5 basado en 1).
    """
    if not path:
        return {}
    try:
        df_coords = pd.read_excel(path, sheet_name="Coordenadas", header=None, dtype=str)
        col_pais, col_uni, skip = _detect_coords_columns(df_coords, default_col_pais, default_col_uni)
        if skip:
            df_coords = df_coords.iloc[1:].reset_index(drop=True)
        if df_coords.shape[1] >= PLAN_ESTUDIOS_COL_IDX + 1:
            unis  = df_coords.iloc[:, col_uni].fillna("").astype(str).str.strip()
            plans = df_coords.iloc[:, PLAN_ESTUDIOS_COL_IDX].fillna("").astype(str).str.strip()
            _bad  = {"nan", "none", ""}
            mask  = unis.ne("") & plans.ne("") & ~plans.str.lower().isin(_bad)
            result = dict(zip(unis[mask], plans[mask]))
            if result:
                return result
        plan_map = _build_plan_estudios_from_students(path)
        if plan_map:
            _write_plan_estudios_to_coordenadas(path, plan_map, col_uni_openpyxl=col_uni + 1)
        return plan_map
    except Exception:
        return {}


def _build_plan_estudios_from_students(path: str) -> dict:
    """Construye {universidad: plan_estudios} escaneando las hojas de datos del Excel."""
    def _is_academic_sheet(name: str) -> bool:
        return bool(re.search(r'\d{4}|\d{2}[-/]\d{2}', str(name)))

    def _pick_col(df: pd.DataFrame, *names: str):
        norm = lambda s: s.strip().lower()
        for n in names:
            for c in df.columns:
                if norm(str(c)) == norm(n):
                    return c
        return None

    plan_map: dict = {}
    try:
        wb_sheets = pd.ExcelFile(path, engine="openpyxl").sheet_names
    except Exception:
        return {}

    hojas = [s for s in wb_sheets if _is_academic_sheet(s)]
    if not hojas:
        hojas = [s for s in wb_sheets if s.lower() != "coordenadas"]

    # Ordenar por curso descendente para priorizar el más reciente
    def _sort_key(s: str):
        m = re.search(r'(\d{4})', str(s))
        return int(m.group(1)) if m else 0
    hojas = sorted(hojas, key=_sort_key, reverse=True)

    for sheet in hojas:
        try:
            df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl", dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
            c_uni  = _pick_col(df, "Destino", "destino", "Universidad", "universidad")
            c_plan = _pick_col(
                df,
                "Plan de estudios", "plan de estudios",
                "Plan estudios", "plan estudios",
                "Enlace plan de estudios", "enlace plan de estudios",
                "Plan_estudios", "plan_estudios",
                "PlanEstudios", "Plan",
            )
            if not c_uni or not c_plan:
                continue
            unis  = df[c_uni].fillna("").astype(str).str.strip()
            plans = df[c_plan].fillna("").astype(str).str.strip()
            _bad  = {"nan", "none", ""}
            mask  = (unis.ne("") & plans.ne("")
                     & ~unis.str.lower().isin(_bad)
                     & ~plans.str.lower().isin(_bad))
            for uni, plan in zip(unis[mask].tolist(), plans[mask].tolist()):
                plan_map.setdefault(uni, plan)
        except Exception:
            continue
    return plan_map


def _write_plan_estudios_to_coordenadas(
    path: str,
    plan_map: dict,
    col_uni_openpyxl: int = 2,
) -> None:
    """Escribe el plan de estudios en col E (col 5, 1-based) de la hoja 'Coordenadas'."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        if "Coordenadas" not in wb.sheetnames:
            return
        ws = wb["Coordenadas"]
        for r_idx in range(1, ws.max_row + 1):
            uni_val = str(ws.cell(row=r_idx, column=col_uni_openpyxl).value or "").strip()
            if uni_val and uni_val.lower() not in ("nan", "none", ""):
                plan_val = plan_map.get(uni_val, "")
                if plan_val:
                    ws.cell(row=r_idx, column=PLAN_ESTUDIOS_COL_XLSX).value = plan_val
        wb.save(path)
        logger.info("Plan de estudios escrito en hoja Coordenadas: %d universidades", len(plan_map))
    except Exception as e:
        logger.warning("No se pudo escribir plan de estudios en Coordenadas: %s", e)




# ──────────────────────────────────────────────────────────────────────────────
# Lista de países (ISO con babel, cacheada)
# ──────────────────────────────────────────────────────────────────────────────

COUNTRY_ALIASES = {
    "Chequia": "República Checa",
    "Eslovaquia": "República Eslovaca",
    "Corea del Sur": "Corea, República de",
    "Corea del Norte": "Corea, República Popular Democrática de",
}


@st.cache_data
def get_country_options() -> list[str]:
    import pycountry
    from babel import Locale
    locale_es = Locale("es")
    nombres = []
    for country in pycountry.countries:
        nombre_es = locale_es.territories.get(country.alpha_2, country.name)
        nombre_es = COUNTRY_ALIASES.get(nombre_es, nombre_es)
        nombres.append(nombre_es)
    return [""] + sorted(set(nombres))


COUNTRY_OPTIONS: list[str] = get_country_options()


# ──────────────────────────────────────────────────────────────────────────────
# Catálogo de asignaturas (con caché en session_state)
# ──────────────────────────────────────────────────────────────────────────────

def load_asignaturas_catalog(config: dict, sheet_name: str | None = None) -> list:
    """Carga el catálogo de asignaturas con caché en session_state.

    La clave incluye data_version para invalidar automáticamente tras
    cualquier guardado (misma lógica que _load_asignaturas_catalog en popup_templates).
    """
    ruta = (config.get("Erasmus IN") or "").strip()
    data_version = st.session_state.get("data_version", 0)
    cache_key = f"_asignaturas_catalog_{ruta}_{sheet_name or ''}_{data_version}"

    if cache_key not in st.session_state:
        try:
            from persistence import get_asignaturas_catalog
            st.session_state[cache_key] = get_asignaturas_catalog(config, sheet_name=sheet_name)
        except Exception as e:
            logger.warning("No se pudo cargar catálogo de asignaturas: %s", e)
            st.session_state[cache_key] = []
    return st.session_state.get(cache_key, [])


# ──────────────────────────────────────────────────────────────────────────────
# Geocodificación (con caché en session_state)
# ──────────────────────────────────────────────────────────────────────────────

try:
    from geopy.geocoders import Nominatim
    from geopy.extra.rate_limiter import RateLimiter
    _GEOCODER = Nominatim(user_agent="tfg-movilidad-esii")
    _GEOCODE = RateLimiter(_GEOCODER.geocode, min_delay_seconds=1)
except Exception:
    _GEOCODER = None
    _GEOCODE = None


def geocode_cached(q: str) -> tuple[float | None, float | None, str | None]:
    if not q or not q.strip():
        return None, None, "Consulta vacía"
    qn = q.strip().lower()
    cache = st.session_state.setdefault("_geo_cache", {})
    if qn in cache:
        d = cache[qn]
        return d["lat"], d["lon"], None
    if _GEOCODE is None:
        return None, None, "Geocoder no disponible"
    try:
        loc = _GEOCODE(q, addressdetails=False, timeout=10)
        if loc:
            lat, lon = float(loc.latitude), float(loc.longitude)
            cache[qn] = {"lat": lat, "lon": lon}
            return lat, lon, None
        return None, None, "No encontrado"
    except Exception as e:
        return None, None, f"Error geocodificando: {e}"


# ──────────────────────────────────────────────────────────────────────────────
# Utilidades de asignaturas y hojas
# ──────────────────────────────────────────────────────────────────────────────

def is_academic_year(name: str) -> bool:
    return bool(re.search(r'\d{4}|\d{2}[-/]\d{2}', name))


def sheet_options_for(cfg: dict, tipo: str) -> list[str]:
    from persistence import sheets_for
    sheets_map = (cfg or {}).get("sheets", {}) or {}
    known = sheets_map.get(tipo) or []
    if known:
        return sorted({s for s in known if s and s != "__CSV__" and is_academic_year(s)})
    path = (cfg or {}).get(tipo, "")
    return [s for s in sheets_for(path) if s != "__CSV__" and is_academic_year(s)] if path else []


def normalize_subject_name(s: str) -> str:
    """Normaliza nombre de asignatura para comparación insensible a acentos/mayúsculas."""
    s = (s or "").strip()
    s = re.sub(r'\s+', ' ', s)
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    return s.casefold()


def asig_nombre_puro(label: str) -> str:
    """Extrae el nombre de asignatura quitando el sufijo '(matr/cupo)'."""
    return label.split("  (")[0].strip() if "  (" in label else label.strip()
